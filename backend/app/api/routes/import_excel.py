"""
Excel → Form Schema importer.

POST /forms/import-excel   (multipart/form-data, field: file)

Returns a json_schema object ready for POST /forms/:
  {
    "title": "<sheet name>",
    "sections": [
      { "title": "1. GENERAL INFORMATION", "fields": [...] },
      ...
    ],
    "fields": [...]          # flat list, same fields with .section set
  }

Algorithm
---------
1. Load workbook (full mode so merged-cell ranges are available).
2. Build a "resolved grid": propagate merge-anchor values to every cell
   in each merge range.
3. Detect where the data rows begin: scan down until we find a row that
   is ≥ 30 % filled AND has ≥ 30 % numeric / date values → that row is
   the first *data* row; everything above is headers.
   Fallback: if no such row found in first 10 rows, pick the last row
   whose fill-ratio exceeds the previous row's.
4. For each column build a hierarchical label by walking the header rows
   top-to-bottom and collecting the *last* (most-specific) non-empty
   resolved value.  Earlier rows give the section path.
5. Sample up to 100 data rows to infer field type.
6. Return structured json_schema.
"""

from __future__ import annotations

import io
import re
import uuid
from datetime import datetime, date
from typing import Any, Optional

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.deps import require_org_admin

router = APIRouter()

# ── helpers ──────────────────────────────────────────────────────────────────

_SKIP_LABELS = {
    "", "na", "n/a", "nil", "-", "–", "—",
    "record no.", "sl.no.", "sl no", "record no",
}

_TYPE_HINTS: list[tuple[re.Pattern, str]] = [
    (re.compile(r"\b(date|month|year|dob|dt)\b", re.I), "date"),
    (re.compile(r"\b(latitude|longitude|lat|lng|long|gps)\b", re.I), "text"),
    (re.compile(r"\b(mobile|phone|contact|whatsapp|no\.?)\b", re.I), "text"),
    (re.compile(r"\b(aadhaar|aadhar|pan|id|code|no\.?|number|serial)\b", re.I), "text"),
    (re.compile(r"\b(name|address|village|town|district|taluk|division|gp|remarks?|note|description|reason|comment)\b", re.I), "text"),
    (re.compile(r"\b(sex|gender|caste|subcaste|occupation|education|source|purpose|type|bank|category|status|pilot|main)\b", re.I), "select"),
    (re.compile(r"\b(amount|rs\.?|income|cost|area|count|members?|children|age|years?|months?|days?|hours?|units?|qty|quantity|yield|score|rating|number of)\b", re.I), "number"),
]


def _clean(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    # strip leading numbers like "1.", "2.1.", "Q3."
    s = re.sub(r"^\s*\d+(\.\d+)*\.\s*", "", s)
    return s if s else None


def _infer_type(label: str, samples: list[Any]) -> str:
    label_l = label.lower()
    # label-based hints first
    for pat, t in _TYPE_HINTS:
        if pat.search(label_l):
            return t

    if not samples:
        return "text"

    # data-based inference
    date_count = num_count = str_count = 0
    unique_vals: set[str] = set()
    for v in samples:
        if isinstance(v, (datetime, date)):
            date_count += 1
        elif isinstance(v, (int, float)):
            num_count += 1
        else:
            s = str(v).strip()
            unique_vals.add(s)
            str_count += 1

    total = len(samples)
    if date_count / total > 0.5:
        return "date"
    if num_count / total > 0.5:
        return "number"
    # low cardinality → select
    if unique_vals and len(unique_vals) <= 8 and str_count / total > 0.5:
        return "select"
    return "text"


def _options_for(samples: list[Any]) -> list[str]:
    """Return sorted unique string values if cardinality ≤ 20."""
    vals = sorted({str(v).strip() for v in samples if v is not None and str(v).strip()})
    if len(vals) <= 20:
        return vals
    return []


# ── resolved grid builder ─────────────────────────────────────────────────────

def _build_resolved_grid(ws) -> dict[tuple[int, int], Any]:
    """
    Return a dict (row, col) → value where every cell inside a merge range
    carries the anchor cell's value.
    """
    grid: dict[tuple[int, int], Any] = {}

    # First fill all actual cell values
    for row in ws.iter_rows():
        for cell in row:
            grid[(cell.row, cell.column)] = cell.value

    # Propagate merged-cell anchors
    for merge in ws.merged_cells.ranges:
        anchor_val = ws.cell(merge.min_row, merge.min_col).value
        for r in range(merge.min_row, merge.max_row + 1):
            for c in range(merge.min_col, merge.max_col + 1):
                grid[(r, c)] = anchor_val

    return grid


# ── header-row detection ──────────────────────────────────────────────────────

def _detect_data_start(grid: dict, max_col: int, max_scan: int = 12) -> int:
    """
    Return the 1-based row number where actual data begins.
    Heuristic: first row that is ≥ 30 % filled AND ≥ 25 % numeric/date values.
    Falls back to the densest row found.
    """
    best_row = 2
    best_fill = 0.0

    for r in range(1, max_scan + 1):
        vals = [grid.get((r, c)) for c in range(1, max_col + 1)]
        non_empty = [v for v in vals if v is not None]
        if not non_empty:
            continue
        fill = len(non_empty) / max_col
        numeric = sum(1 for v in non_empty if isinstance(v, (int, float, datetime, date)))
        num_ratio = numeric / len(non_empty)

        if fill >= 0.3 and num_ratio >= 0.25:
            return r          # first data-like row

        if fill > best_fill:
            best_fill = fill
            best_row = r

    # No clear data row → assume headers are rows 1..(best_row-1)
    return best_row + 1


# ── main parser ───────────────────────────────────────────────────────────────

def parse_excel_to_schema(file_bytes: bytes, title_hint: str = "") -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    sheet_title = ws.title or title_hint or "Imported Form"

    max_col = ws.max_column or 1
    max_row = ws.max_row or 1

    grid = _build_resolved_grid(ws)
    data_start = _detect_data_start(grid, max_col)
    header_rows = list(range(1, data_start))   # e.g. [1,2,3,4,5]

    if not header_rows:
        header_rows = [1]
        data_start = 2

    # ── build per-column label hierarchy ─────────────────────────────────────
    fields: list[dict] = []

    for c in range(1, max_col + 1):
        # Collect non-empty resolved values from each header row
        parts: list[str] = []
        for r in header_rows:
            raw = grid.get((r, c))
            v = _clean(raw)
            if v and v.lower() not in _SKIP_LABELS:
                # Only append if different from the last part (avoid "Section > Section")
                if not parts or v != parts[-1]:
                    parts.append(v)

        if not parts:
            continue

        # Most specific label = last part; section = everything above it
        label = parts[-1]
        section_path = parts[:-1]

        # Skip boilerplate columns
        if label.lower() in _SKIP_LABELS:
            continue

        # Sample data for type inference
        samples: list[Any] = []
        for r in range(data_start, min(data_start + 100, max_row + 1)):
            v = grid.get((r, c))
            if v is not None:
                samples.append(v)

        ftype = _infer_type(label, samples)

        field: dict[str, Any] = {
            "id": f"f{c}",
            "label": label,
            "type": ftype,
            "required": False,
            "section": " > ".join(section_path) if section_path else "",
        }

        if ftype == "select":
            opts = _options_for(samples)
            if opts:
                field["options"] = opts

        fields.append(field)

    # ── group into sections ────────────────────────────────────────────────────
    # Use the top-level part of section_path as the section title
    section_map: dict[str, list[dict]] = {}
    for f in fields:
        sec = f.get("section") or "General"
        # Use first segment for grouping (before the first " > ")
        top = sec.split(" > ")[0].strip() if sec else "General"
        section_map.setdefault(top, []).append(f)

    sections = [
        {"title": title, "fields": flds}
        for title, flds in section_map.items()
    ]

    wb.close()

    return {
        "title": sheet_title,
        "sections": sections,
        "fields": fields,
    }


# ── route ─────────────────────────────────────────────────────────────────────

@router.post("/import-excel")
async def import_excel(
    file: UploadFile = File(...),
    user=Depends(require_org_admin),
    db: Session = Depends(get_db),
):
    """
    Parse an Excel (.xlsx) file and return a draft json_schema.
    The caller can review / edit the schema then POST it to /forms/.
    """
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx / .xls files are accepted")

    content = await file.read()
    if len(content) > 20 * 1024 * 1024:  # 20 MB guard
        raise HTTPException(status_code=413, detail="File too large (max 20 MB)")

    try:
        schema = parse_excel_to_schema(content, title_hint=file.filename.rsplit(".", 1)[0])
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Could not parse Excel file: {e}")

    return schema
