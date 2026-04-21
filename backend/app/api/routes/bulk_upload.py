"""
Bulk Upload — Excel/CSV → Submissions (org_admin and master_admin only)

Flow
----
1. POST /bulk-upload/parse
       multipart: file (.xlsx/.csv) + form_id
   → returns detected columns, form fields, suggested mapping, preview rows,
     and an upload_token (base64-gzipped JSON of all rows so no server state needed)

2. POST /bulk-upload/apply
       JSON: {upload_token, form_id, mapping, meta_lat, meta_lng,
              meta_datetime, meta_enumerator}
   → creates submissions; returns {created, skipped, errors}

3. GET  /bulk-upload/template/{form_id}
   → streams an Excel template (.xlsx) pre-filled with form headers,
     type hints, example row, and optional metadata columns
"""
from __future__ import annotations

import base64
import gzip
import io
import json
import re
import uuid
from datetime import datetime, timezone
from difflib import SequenceMatcher
from typing import Any, Optional

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl.styles import (Alignment, Font, PatternFill, Border, Side)
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.deps import get_current_user
from app.models.form import Form as FormModel
from app.models.submission import Submission
from app.models.user import User

router = APIRouter()

# ── Auth guard: org_admin OR master_admin ─────────────────────────────────────

def require_bulk_admin(user=Depends(get_current_user)):
    if user.get("role") not in ("org_admin", "master_admin"):
        raise HTTPException(status_code=403, detail="Bulk upload requires org_admin or master_admin role")
    return user

# ── Section colours (cycling) ─────────────────────────────────────────────────

_SEC_COLORS = [
    "BBDEFB",  # blue-100
    "C8E6C9",  # green-100
    "FFF9C4",  # yellow-100
    "F8BBD0",  # pink-100
    "E1BEE7",  # purple-100
    "FFE0B2",  # orange-100
]
_META_COLOR = "CFD8DC"  # blue-grey-100

# ── helpers ───────────────────────────────────────────────────────────────────

def _extract_fields(form: FormModel) -> list[dict]:
    """Flatten form fields from sections + top-level fields list."""
    schema = form.json_schema or {}
    fields: list[dict] = []
    seen_ids: set[str] = set()

    def _add(f: dict):
        fid = f.get("id") or f.get("key") or f.get("name")
        if fid and fid not in seen_ids:
            seen_ids.add(fid)
            fields.append({
                "id": fid,
                "label": f.get("label", fid),
                "type": f.get("type", "text"),
                "required": f.get("required", False),
                "options": f.get("options", []),
                "section": f.get("section", ""),
            })

    for sec in schema.get("sections", []):
        for f in sec.get("fields", []):
            if "section" not in f:
                f = dict(f, section=sec.get("title", ""))
            _add(f)

    for f in schema.get("fields", []):
        _add(f)

    return fields


def _similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


def _suggest_mapping(form_fields: list[dict], excel_cols: list[str]) -> dict[str, str]:
    """Return {field_id: excel_column} best-guess mapping."""
    mapping: dict[str, str] = {}
    used_cols: set[str] = set()

    for field in form_fields:
        label = field["label"]
        best_col, best_score = None, 0.0
        for col in excel_cols:
            s = _similarity(label, col)
            if s > best_score:
                best_score, best_col = s, col
        if best_col and best_score >= 0.45 and best_col not in used_cols:
            mapping[field["id"]] = best_col
            used_cols.add(best_col)

    return mapping


def _suggest_meta(excel_cols: list[str]) -> dict[str, Optional[str]]:
    """Auto-detect latitude, longitude, datetime, enumerator columns."""
    lat_pat = re.compile(r"\b(lat(itude)?)\b", re.I)
    lng_pat = re.compile(r"\b(lon(g(itude)?)?|lng)\b", re.I)
    dt_pat  = re.compile(r"\b(date|datetime|collected.?at|timestamp|time)\b", re.I)
    en_pat  = re.compile(r"\b(enumerator|surveyor|field.?worker|collector|phone)\b", re.I)

    meta: dict[str, Optional[str]] = {
        "meta_lat": None, "meta_lng": None,
        "meta_datetime": None, "meta_enumerator": None,
    }
    for col in excel_cols:
        if lat_pat.search(col) and not meta["meta_lat"]:
            meta["meta_lat"] = col
        if lng_pat.search(col) and not meta["meta_lng"]:
            meta["meta_lng"] = col
        if dt_pat.search(col) and not meta["meta_datetime"]:
            meta["meta_datetime"] = col
        if en_pat.search(col) and not meta["meta_enumerator"]:
            meta["meta_enumerator"] = col
    return meta


def _parse_file(content: bytes, filename: str) -> tuple[list[str], list[dict]]:
    """Parse xlsx or csv; return (column_names, rows_as_dicts)."""
    if filename.lower().endswith(".csv"):
        import csv as _csv
        text = content.decode("utf-8-sig", errors="replace")
        reader = _csv.DictReader(io.StringIO(text))
        cols = reader.fieldnames or []
        rows = [dict(r) for r in reader]
        return list(cols), rows

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    # Resolve merged cells
    grid: dict[tuple[int, int], Any] = {}
    for row in ws.iter_rows():
        for cell in row:
            grid[(cell.row, cell.column)] = cell.value
    for merge in ws.merged_cells.ranges:
        anchor = ws.cell(merge.min_row, merge.min_col).value
        for r in range(merge.min_row, merge.max_row + 1):
            for c in range(merge.min_col, merge.max_col + 1):
                grid[(r, c)] = anchor

    if ws.max_row < 1:
        return [], []

    # Row 1 = headers
    max_col = ws.max_column or 1
    cols: list[str] = []
    for c in range(1, max_col + 1):
        v = grid.get((1, c))
        cols.append(str(v).strip() if v is not None else f"Column_{c}")

    # Remaining rows = data
    rows: list[dict] = []
    for r in range(2, ws.max_row + 1):
        row_dict = {}
        non_empty = False
        for c_idx, col in enumerate(cols, start=1):
            v = grid.get((r, c_idx))
            if v is not None:
                non_empty = True
            row_dict[col] = v
        if non_empty:
            rows.append(row_dict)

    wb.close()
    return cols, rows


def _encode_token(rows: list[dict]) -> str:
    raw = json.dumps(rows, default=str).encode()
    return base64.b64encode(gzip.compress(raw)).decode()


def _decode_token(token: str) -> list[dict]:
    try:
        raw = gzip.decompress(base64.b64decode(token.encode()))
        return json.loads(raw)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid upload token")


def _parse_datetime(v: Any) -> Optional[datetime]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.replace(tzinfo=timezone.utc) if v.tzinfo is None else v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
                "%d-%m-%Y %H:%M", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            pass
    return None


def _parse_float(v: Any) -> Optional[float]:
    try:
        return float(str(v).strip())
    except (TypeError, ValueError):
        return None


# ── PARSE ─────────────────────────────────────────────────────────────────────

@router.post("/parse")
async def bulk_upload_parse(
    form_id: str = Form(...),
    file: UploadFile = File(...),
    user=Depends(require_bulk_admin),
    db: Session = Depends(get_db),
):
    """Step 1 — Parse uploaded file and return column-mapping info."""
    fname = file.filename or ""
    if not fname.lower().endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(status_code=400, detail="Only .xlsx, .xls, or .csv files accepted")

    content = await file.read()
    if len(content) > 50 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (max 50 MB)")

    # Load form — master_admin can access any tenant's form via form_id
    if user["role"] == "master_admin":
        form = db.query(FormModel).filter(FormModel.id == form_id).first()
    else:
        form = db.query(FormModel).filter(
            FormModel.id == form_id,
            FormModel.tenant_id == user["tenant_id"],
        ).first()

    if not form:
        raise HTTPException(status_code=404, detail="Form not found")

    form_fields = _extract_fields(form)

    try:
        excel_cols, rows = _parse_file(content, fname)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Could not parse file: {e}")

    if not rows:
        raise HTTPException(status_code=422, detail="File has no data rows")

    token = _encode_token(rows)
    suggested = _suggest_mapping(form_fields, excel_cols)
    meta = _suggest_meta(excel_cols)

    preview = []
    for row in rows[:5]:
        preview.append({k: (str(v) if v is not None else None) for k, v in row.items()})

    return {
        "form_id": str(form.id),
        "form_title": form.title,
        "form_fields": form_fields,
        "excel_columns": excel_cols,
        "row_count": len(rows),
        "suggested_mapping": suggested,
        "suggested_meta": meta,
        "preview": preview,
        "upload_token": token,
    }


# ── APPLY ─────────────────────────────────────────────────────────────────────

class ApplyRequest(BaseModel):
    upload_token: str
    form_id: str
    mapping: dict[str, str]          # {field_id: excel_column}
    meta_lat: Optional[str] = None
    meta_lng: Optional[str] = None
    meta_datetime: Optional[str] = None
    meta_enumerator: Optional[str] = None


@router.post("/apply")
def bulk_upload_apply(
    body: ApplyRequest,
    user=Depends(require_bulk_admin),
    db: Session = Depends(get_db),
):
    """Step 2 — Apply column mapping and create submissions."""
    if user["role"] == "master_admin":
        form = db.query(FormModel).filter(FormModel.id == body.form_id).first()
    else:
        form = db.query(FormModel).filter(
            FormModel.id == body.form_id,
            FormModel.tenant_id == user["tenant_id"],
        ).first()

    if not form:
        raise HTTPException(status_code=404, detail="Form not found")

    tenant_id = str(form.tenant_id)
    rows = _decode_token(body.upload_token)

    # Cache enumerator lookups {phone_or_name: user_id}
    _enum_cache: dict[str, Any] = {}

    def _resolve_enumerator(val: Any) -> Optional[str]:
        if val is None:
            return None
        key = str(val).strip()
        if key in _enum_cache:
            return _enum_cache[key]
        # Try by phone, then by name
        u = (
            db.query(User).filter(User.tenant_id == tenant_id, User.phone == key).first()
            or db.query(User).filter(User.tenant_id == tenant_id, User.name.ilike(key)).first()
        )
        result = str(u.id) if u else None
        _enum_cache[key] = result
        return result

    uploader_id = user["sub"]
    created = 0
    skipped = 0
    errors: list[str] = []

    for i, row in enumerate(rows, start=2):
        try:
            data_json: dict[str, Any] = {}
            for field_id, col_name in body.mapping.items():
                if col_name and col_name in row:
                    v = row[col_name]
                    data_json[field_id] = v if v is not None else None

            if not data_json:
                skipped += 1
                continue

            # GPS from metadata columns
            gps: Optional[dict] = None
            lat = _parse_float(row.get(body.meta_lat)) if body.meta_lat else None
            lng = _parse_float(row.get(body.meta_lng)) if body.meta_lng else None
            if lat is not None and lng is not None:
                gps = {"lat": lat, "lng": lng, "accuracy": 0, "source": "bulk_upload"}

            # Collection datetime
            collected_at: Optional[datetime] = None
            if body.meta_datetime:
                collected_at = _parse_datetime(row.get(body.meta_datetime))

            # Enumerator
            enum_id_str: Optional[str] = None
            if body.meta_enumerator:
                enum_id_str = _resolve_enumerator(row.get(body.meta_enumerator))
            if not enum_id_str:
                enum_id_str = uploader_id

            now = datetime.now(timezone.utc)
            sub = Submission(
                id=uuid.uuid4(),
                tenant_id=tenant_id,
                form_id=form.id,
                form_version=form.version,
                enumerator_id=enum_id_str,
                local_id=str(uuid.uuid4()),
                data_json=data_json,
                gps_open=gps,
                gps_submit=gps,
                status="synced",
                local_created_at=collected_at or now,
                server_received_at=now,
            )
            db.add(sub)
            created += 1

        except Exception as e:
            errors.append(f"Row {i}: {e}")
            skipped += 1

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {e}")

    return {
        "created": created,
        "skipped": skipped,
        "errors": errors[:20],
        "form_id": str(form.id),
        "form_title": form.title,
    }


# ── TEMPLATE DOWNLOAD ─────────────────────────────────────────────────────────

@router.get("/template/{form_id}")
def download_template(
    form_id: str,
    user=Depends(require_bulk_admin),
    db: Session = Depends(get_db),
):
    """Download a pre-filled Excel template for a form."""
    if user["role"] == "master_admin":
        form = db.query(FormModel).filter(FormModel.id == form_id).first()
    else:
        form = db.query(FormModel).filter(
            FormModel.id == form_id,
            FormModel.tenant_id == user["tenant_id"],
        ).first()

    if not form:
        raise HTTPException(status_code=404, detail="Form not found")

    fields = _extract_fields(form)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = form.title[:31]  # Excel sheet name limit

    # ── group fields by section ───────────────────────────────────────────────
    sections: dict[str, list[dict]] = {}
    for f in fields:
        sec = f.get("section") or "General"
        sections.setdefault(sec, []).append(f)

    # ── styles ────────────────────────────────────────────────────────────────
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    def _hdr_font(bold=False, italic=False, color="000000") -> Font:
        return Font(bold=bold, italic=italic, color=color, size=10, name="Calibri")

    # ── Row 1: Section headers (merged, colored) ──────────────────────────────
    col = 1
    sec_start_cols: list[tuple[str, int, int, str]] = []
    for sec_idx, (sec_name, sec_fields) in enumerate(sections.items()):
        start = col
        end = col + len(sec_fields) - 1
        color = _SEC_COLORS[sec_idx % len(_SEC_COLORS)]
        sec_start_cols.append((sec_name, start, end, color))
        col = end + 1

    # Metadata columns at the end
    meta_cols = ["_Latitude", "_Longitude", "_Collected At (YYYY-MM-DD HH:MM)", "_Enumerator Phone"]
    meta_start = col
    meta_end = col + len(meta_cols) - 1

    # Write section header row
    for sec_name, start, end, color in sec_start_cols:
        cell = ws.cell(row=1, column=start, value=sec_name)
        cell.font = _hdr_font(bold=True, color="1A237E")
        cell.fill = _fill(color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        if end > start:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
            for c in range(start + 1, end + 1):
                ws.cell(row=1, column=c).fill = _fill(color)
                ws.cell(row=1, column=c).border = border

    # Metadata section header
    meta_hdr = ws.cell(row=1, column=meta_start, value="★ Metadata (optional — improves data quality)")
    meta_hdr.font = _hdr_font(bold=True, color="37474F")
    meta_hdr.fill = _fill(_META_COLOR)
    meta_hdr.alignment = Alignment(horizontal="center", vertical="center")
    meta_hdr.border = border
    if meta_end > meta_start:
        ws.merge_cells(start_row=1, start_column=meta_start, end_row=1, end_column=meta_end)
        for c in range(meta_start + 1, meta_end + 1):
            ws.cell(row=1, column=c).fill = _fill(_META_COLOR)
            ws.cell(row=1, column=c).border = border

    # ── Row 2: Field labels ───────────────────────────────────────────────────
    col = 1
    for sec_name, sec_fields, color in [
        (sn, sf, c) for sn, start, end, c in sec_start_cols
        for sf in [sections[sn]]
    ]:
        for f in sec_fields:
            label = f["label"]
            req = f["required"]
            cell = ws.cell(row=2, column=col, value=("* " + label) if req else label)
            cell.font = _hdr_font(bold=True, color=("C62828" if req else "212121"))
            cell.fill = _fill(color)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = border
            col += 1

    for mc in meta_cols:
        cell = ws.cell(row=2, column=col, value=mc)
        cell.font = _hdr_font(bold=True, color="455A64")
        cell.fill = _fill(_META_COLOR)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border
        col += 1

    # ── Row 3: Type hints ─────────────────────────────────────────────────────
    col = 1
    for sec_name, sec_fields, color in [
        (sn, sf, c) for sn, start, end, c in sec_start_cols
        for sf in [sections[sn]]
    ]:
        for f in sec_fields:
            ftype = f["type"]
            opts = f.get("options", [])
            if opts:
                hint = "Select: " + " / ".join(opts[:5]) + ("…" if len(opts) > 5 else "")
            elif ftype == "gps":
                hint = "Leave blank — use _Latitude/_Longitude columns"
            elif ftype == "photo":
                hint = "Leave blank — photos captured in app"
            elif ftype == "date":
                hint = "YYYY-MM-DD"
            elif ftype == "number":
                hint = "Numeric only"
            else:
                hint = ftype
            cell = ws.cell(row=3, column=col, value=hint)
            cell.font = _hdr_font(italic=True, color="757575")
            cell.fill = _fill("FAFAFA")
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = border
            col += 1

    meta_hints = [
        "Decimal degrees e.g. 12.9716",
        "Decimal degrees e.g. 77.5946",
        "2025-06-15 10:30  or  2025-06-15",
        "Phone number or enumerator name",
    ]
    for hint in meta_hints:
        cell = ws.cell(row=3, column=col, value=hint)
        cell.font = _hdr_font(italic=True, color="78909C")
        cell.fill = _fill("F5F5F5")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border
        col += 1

    # ── Row 4: Example row ────────────────────────────────────────────────────
    _TYPE_EXAMPLES: dict[str, str] = {
        "text": "Example text",
        "number": "42",
        "select": "— see options above —",
        "date": "2025-06-15",
        "gps": "",
        "photo": "",
        "audio": "",
        "textarea": "Free text notes here",
        "barcode": "1234567890",
    }
    col = 1
    for sec_name, sec_fields, color in [
        (sn, sf, c) for sn, start, end, c in sec_start_cols
        for sf in [sections[sn]]
    ]:
        for f in sec_fields:
            opts = f.get("options", [])
            ex = opts[0] if opts else _TYPE_EXAMPLES.get(f["type"], "")
            cell = ws.cell(row=4, column=col, value=ex)
            cell.font = _hdr_font(italic=True, color="388E3C")
            cell.fill = _fill("F1F8E9")
            cell.border = border
            col += 1

    meta_ex = ["12.9716", "77.5946", "2025-06-15 10:30", "+919999990003"]
    for ex in meta_ex:
        cell = ws.cell(row=4, column=col, value=ex)
        cell.font = _hdr_font(italic=True, color="546E7A")
        cell.fill = _fill("ECEFF1")
        cell.border = border
        col += 1

    # ── Add data validation for select fields ──────────────────────────────────
    from openpyxl.worksheet.datavalidation import DataValidation
    col = 1
    for sec_name, sec_fields, color in [
        (sn, sf, c) for sn, start, end, c in sec_start_cols
        for sf in [sections[sn]]
    ]:
        for f in sec_fields:
            opts = f.get("options", [])
            if opts and len(opts) <= 50:
                formula = '"' + ",".join(str(o) for o in opts[:50]) + '"'
                col_letter = get_column_letter(col)
                dv = DataValidation(
                    type="list",
                    formula1=formula,
                    allow_blank=True,
                    showDropDown=False,
                )
                dv.sqref = f"{col_letter}5:{col_letter}10000"
                ws.add_data_validation(dv)
            col += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    total_cols = meta_end
    col = 1
    all_fields_flat = [f for fields_list in sections.values() for f in fields_list]
    for f in all_fields_flat:
        label_len = max(len(f["label"]) + 2, 14)
        ws.column_dimensions[get_column_letter(col)].width = min(label_len, 28)
        col += 1
    meta_widths = [14, 14, 30, 22]
    for w in meta_widths:
        ws.column_dimensions[get_column_letter(col)].width = w
        col += 1

    # Row heights
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 36
    ws.row_dimensions[4].height = 18

    # Freeze panes below row 4, after col 0
    ws.freeze_panes = "A5"

    # ── Instruction sheet ─────────────────────────────────────────────────────
    ws_info = wb.create_sheet("Instructions")
    instructions = [
        ["FieldGovern Bulk Upload — Instructions"],
        [""],
        ["COLUMN LEGEND"],
        ["* (asterisk) = required field — must not be blank"],
        ["Row 2 = field labels (use these as column headers in your data)"],
        ["Row 3 = type hints and allowed values for select fields"],
        ["Row 4 = example row (delete before uploading)"],
        [""],
        ["METADATA COLUMNS (optional but recommended)"],
        ["_Latitude         Decimal degrees (e.g. 12.9716). GPS location of survey."],
        ["_Longitude        Decimal degrees (e.g. 77.5946)."],
        ["_Collected At     Date/time the data was actually collected. Format: YYYY-MM-DD HH:MM"],
        ["                  If omitted, upload timestamp is used."],
        ["_Enumerator Phone Phone number or name of the field worker who collected the data."],
        ["                  Must match a user in the system. If blank, uploader is used."],
        [""],
        ["HOW TO UPLOAD"],
        ["1. Fill data starting from row 5 (delete the example row 4 if you wish)."],
        ["2. Login as Org Admin or Master Admin → Data → Bulk Upload."],
        ["3. Select the form, upload this file."],
        ["4. Map columns to form fields in the mapping step."],
        ["5. Confirm — data appears in the dashboard immediately."],
        [""],
        ["TIPS"],
        ["• Do NOT modify the header rows (rows 1-3)."],
        ["• Select fields: only values from the dropdown list are valid."],
        ["• GPS fields in the form are skipped — use the metadata columns instead."],
        ["• Photo/Audio fields cannot be bulk uploaded; they are collected in-app."],
        ["• Maximum 10,000 rows per upload (split larger files into batches)."],
    ]
    for row_data in instructions:
        ws_info.append(row_data)

    ws_info.column_dimensions["A"].width = 80
    ws_info["A1"].font = Font(bold=True, size=14, color="0D47A1")
    ws_info["A3"].font = Font(bold=True, size=11)
    ws_info["A9"].font = Font(bold=True, size=11)
    ws_info["A17"].font = Font(bold=True, size=11)
    ws_info["A23"].font = Font(bold=True, size=11)

    # ── Stream response ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_title = re.sub(r"[^\w\s-]", "", form.title).strip().replace(" ", "_")
    filename = f"FieldGovern_Template_{safe_title}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
