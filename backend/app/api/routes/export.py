"""Export endpoints — CSV, Stata .dta, and Google Sheets for form submissions."""

import csv
import io
from datetime import datetime
from typing import Optional

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse, JSONResponse
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.deps import require_supervisor
from app.core.security import verify_api_key
from app.models.api_key import ApiKey
from app.models.form import Form
from app.models.submission import Submission
from app.models.user import User
from fastapi import Query as _Query

router = APIRouter()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _flatten(data: dict, parent_key: str = "", sep: str = "__") -> dict:
    """Recursively flatten nested dicts/lists into dot-separated keys."""
    items: dict = {}
    for k, v in data.items():
        key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.update(_flatten(v, key, sep))
        elif isinstance(v, list):
            for i, elem in enumerate(v):
                if isinstance(elem, dict):
                    items.update(_flatten(elem, f"{key}__{i}", sep))
                else:
                    items[f"{key}__{i}"] = elem
        else:
            items[key] = v
    return items


def _query_submissions(
    db: Session,
    form_id: str,
    tenant_id: str,
    date_from: Optional[datetime],
    date_to: Optional[datetime],
    status: Optional[str],
):
    """Build a filtered query for submissions."""
    q = db.query(Submission).filter(
        Submission.form_id == form_id,
        Submission.tenant_id == tenant_id,
    )
    if date_from:
        q = q.filter(Submission.local_created_at >= date_from)
    if date_to:
        q = q.filter(Submission.local_created_at <= date_to)
    if status:
        q = q.filter(Submission.status == status)
    return q.all()


def _build_enumerator_map(db: Session, tenant_id: str) -> dict:
    """Return {user_id_uuid: name} for all users in the tenant."""
    users = db.query(User).filter(User.tenant_id == tenant_id).all()
    return {u.id: u.name or u.phone for u in users}


def _build_rows(subs: list, enumerator_map: dict) -> list[dict]:
    """Convert submission ORM objects into flat dicts for export."""
    rows = []
    for s in subs:
        flat = _flatten(s.data_json or {})
        row = {
            "submission_id": str(s.id),
            "enumerator_id": str(s.enumerator_id) if s.enumerator_id else "",
            "enumerator_name": enumerator_map.get(s.enumerator_id, ""),
            "form_version": s.form_version or "",
            "status": s.status or "",
            "gps_open_lat": (s.gps_open or {}).get("lat"),
            "gps_open_lng": (s.gps_open or {}).get("lng"),
            "gps_open_accuracy": (s.gps_open or {}).get("accuracy"),
            "gps_submit_lat": (s.gps_submit or {}).get("lat"),
            "gps_submit_lng": (s.gps_submit or {}).get("lng"),
            "gps_submit_accuracy": (s.gps_submit or {}).get("accuracy"),
            "local_created_at": str(s.local_created_at) if s.local_created_at else "",
            "server_received_at": str(s.server_received_at) if s.server_received_at else "",
            **flat,
        }
        rows.append(row)
    return rows


def _sanitize_stata_colname(name: str) -> str:
    """Stata variable names: max 32 chars, alphanumeric + underscore, start with letter/underscore."""
    clean = "".join(c if c.isalnum() or c == "_" else "_" for c in name)
    if clean and clean[0].isdigit():
        clean = "_" + clean
    return clean[:32]


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------

@router.get("/{form_id}/csv")
def export_csv(
    form_id: str,
    date_from: Optional[datetime] = Query(None, description="Filter: local_created_at >="),
    date_to: Optional[datetime] = Query(None, description="Filter: local_created_at <="),
    status: Optional[str] = Query(None, description="Filter by submission status"),
    user: dict = Depends(require_supervisor),
    db: Session = Depends(get_db),
):
    """Export submissions for a form as CSV."""
    form = db.query(Form).filter(
        Form.id == form_id, Form.tenant_id == user["tenant_id"]
    ).first()
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")

    subs = _query_submissions(db, form_id, user["tenant_id"], date_from, date_to, status)
    enum_map = _build_enumerator_map(db, user["tenant_id"])
    rows = _build_rows(subs, enum_map)

    if not rows:
        return StreamingResponse(
            iter([""]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{form.title}.csv"'},
        )

    # Union all keys across rows so every column appears
    all_keys: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for k in row:
            if k not in seen:
                all_keys.append(k)
                seen.add(k)

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=all_keys, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    buf.seek(0)

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{form.title}.csv"'},
    )


# ---------------------------------------------------------------------------
# JSON export (API key or JWT auth)
# ---------------------------------------------------------------------------

def _resolve_user_from_request(request: Request, db: Session) -> dict:
    """Resolve user from X-API-Key header or fall back to JWT Bearer."""
    api_key_header = request.headers.get("X-API-Key")
    if api_key_header:
        records = db.query(ApiKey).filter(ApiKey.is_active == True).all()
        matching = None
        for record in records:
            if verify_api_key(api_key_header, record.key_hash):
                matching = record
                break
        if not matching:
            raise HTTPException(status_code=401, detail="Invalid API key")
        from sqlalchemy import func
        matching.last_used_at = func.now()
        db.commit()
        creator = db.query(User).filter(User.id == matching.created_by_id).first()
        if not creator:
            raise HTTPException(status_code=401, detail="Key creator no longer exists")
        if creator.role not in ("org_admin", "supervisor"):
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        from app.core.database import set_tenant_context
        set_tenant_context(db, str(matching.tenant_id))
        return {
            "sub": str(creator.id),
            "tenant_id": str(matching.tenant_id),
            "role": creator.role,
            "api_key_id": str(matching.id),
        }
    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Authentication required")
    from jose import JWTError
    from app.core.security import decode_token
    from app.core.database import set_tenant_context
    try:
        payload = decode_token(auth_header[7:])
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")
    if payload.get("role") not in ("org_admin", "supervisor"):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    set_tenant_context(db, payload["tenant_id"])
    return payload


@router.get("/submissions/{form_id}/json")
def export_json(
    form_id: str,
    request: Request,
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Flat JSON array of all submissions. Designed for external tools (FG Analyzer, FG Data Cleaner) to consume via API key auth."""
    user = _resolve_user_from_request(request, db)

    form = db.query(Form).filter(
        Form.id == form_id, Form.tenant_id == user["tenant_id"]
    ).first()
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")

    date_from_dt: Optional[datetime] = None
    date_to_dt: Optional[datetime] = None
    if date_from:
        try:
            date_from_dt = datetime.fromisoformat(date_from)
        except ValueError:
            raise HTTPException(status_code=422, detail="Invalid date_from format")
    if date_to:
        try:
            date_to_dt = datetime.fromisoformat(date_to)
        except ValueError:
            raise HTTPException(status_code=422, detail="Invalid date_to format")

    subs = _query_submissions(db, form_id, user["tenant_id"], date_from_dt, date_to_dt, status)
    enum_map = _build_enumerator_map(db, user["tenant_id"])

    result = []
    meta = {
        "_meta": True,
        "form_title": form.title,
        "field_count": len([
            f for section in (form.json_schema or {}).get("sections", [])
            for f in section.get("fields", [])
        ]),
        "submission_count": len(subs),
        "exported_at": datetime.utcnow().isoformat() + "Z",
    }
    result.append(meta)

    for s in subs:
        flat = _flatten(s.data_json or {})
        row = {
            "submission_id": str(s.id),
            "serial_no": s.serial_no,
            "enumerator_name": enum_map.get(s.enumerator_id, ""),
            "submitted_at": s.local_created_at.isoformat() if s.local_created_at else (
                s.server_received_at.isoformat() if s.server_received_at else None
            ),
            "status": s.status or "",
            **flat,
        }
        result.append(row)

    return JSONResponse(content=result)


# ---------------------------------------------------------------------------
# PDF summary report
# ---------------------------------------------------------------------------

@router.get("/{form_id}/pdf")
def export_pdf(
    form_id: str,
    date_from: Optional[datetime] = Query(None),
    date_to: Optional[datetime] = Query(None),
    status: Optional[str] = Query(None),
    user: dict = Depends(require_supervisor),
    db: Session = Depends(get_db),
):
    """Generate a PDF summary report for a form."""
    try:
        from fpdf import FPDF
    except ImportError:
        raise HTTPException(status_code=501, detail="fpdf2 not installed — run: pip install fpdf2")

    form = db.query(Form).filter(Form.id == form_id, Form.tenant_id == user["tenant_id"]).first()
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")

    subs = _query_submissions(db, form_id, user["tenant_id"], date_from, date_to, status)
    enum_map = _build_enumerator_map(db, user["tenant_id"])

    total = len(subs)
    status_counts: dict = {}
    enum_stats: dict = {}

    for s in subs:
        status_counts[s.status] = status_counts.get(s.status, 0) + 1
        ename = enum_map.get(s.enumerator_id, "Unknown")
        if ename not in enum_stats:
            enum_stats[ename] = {"total": 0, "approved": 0, "flagged": 0, "rejected": 0, "synced": 0}
        enum_stats[ename]["total"] += 1
        enum_stats[ename][s.status] = enum_stats[ename].get(s.status, 0) + 1

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Title block
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(14, 165, 233)
    pdf.cell(0, 12, "FieldGovern Data Report", ln=True)
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_text_color(30, 30, 46)
    title = form.title[:80]
    pdf.cell(0, 8, title, ln=True)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(100, 100, 120)
    pdf.cell(0, 6, f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", ln=True)
    if date_from:
        to_str = date_to.strftime("%Y-%m-%d") if date_to else "present"
        pdf.cell(0, 6, f"Period: {date_from.strftime('%Y-%m-%d')} to {to_str}", ln=True)
    pdf.ln(3)
    pdf.set_draw_color(14, 165, 233)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(4)

    # Summary
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(30, 30, 46)
    pdf.cell(0, 7, "Summary", ln=True)
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(60, 60, 80)
    pdf.cell(60, 6, "Total Submissions:")
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, str(total), ln=True)
    pdf.set_font("Helvetica", "", 10)
    for st in ["approved", "synced", "flagged", "rejected"]:
        cnt = status_counts.get(st, 0)
        if cnt > 0:
            pdf.cell(60, 6, f"  {st.capitalize()}:")
            pdf.cell(0, 6, str(cnt), ln=True)
    pdf.ln(4)

    # Enumerator performance table
    if enum_stats:
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(30, 30, 46)
        pdf.cell(0, 7, "Enumerator Performance", ln=True)
        pdf.ln(1)

        col_w = [72, 22, 25, 22, 22]
        headers = ["Enumerator", "Total", "Approved", "Flagged", "Rejected"]
        pdf.set_fill_color(14, 165, 233)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 9)
        for i, h in enumerate(headers):
            pdf.cell(col_w[i], 7, h, border=0, align="C" if i > 0 else "L", fill=True)
        pdf.ln()

        pdf.set_text_color(30, 30, 46)
        pdf.set_font("Helvetica", "", 9)
        for idx, (ename, st) in enumerate(sorted(enum_stats.items(), key=lambda x: -x[1]["total"])):
            fill = idx % 2 == 0
            if fill:
                pdf.set_fill_color(235, 245, 255)
            pdf.cell(col_w[0], 6, ename[:42], border=0, fill=fill)
            pdf.cell(col_w[1], 6, str(st["total"]), border=0, align="C", fill=fill)
            pdf.cell(col_w[2], 6, str(st.get("approved", 0)), border=0, align="C", fill=fill)
            pdf.cell(col_w[3], 6, str(st.get("flagged", 0)), border=0, align="C", fill=fill)
            pdf.cell(col_w[4], 6, str(st.get("rejected", 0)), border=0, align="C", fill=fill)
            pdf.ln()

    buf = io.BytesIO()
    pdf.output(buf)
    buf.seek(0)
    safe_title = "".join(c for c in form.title if c.isalnum() or c in " _-")

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{safe_title}_report.pdf"'},
    )


# ---------------------------------------------------------------------------
# Stata .dta export
# ---------------------------------------------------------------------------

@router.get("/{form_id}/dta")
def export_dta(
    form_id: str,
    date_from: Optional[datetime] = Query(None, description="Filter: local_created_at >="),
    date_to: Optional[datetime] = Query(None, description="Filter: local_created_at <="),
    status: Optional[str] = Query(None, description="Filter by submission status"),
    user: dict = Depends(require_supervisor),
    db: Session = Depends(get_db),
):
    """Export submissions for a form as Stata .dta file."""
    if not HAS_PANDAS:
        raise HTTPException(status_code=501, detail="Stata export requires pandas. Install pandas to enable this feature.")

    form = db.query(Form).filter(
        Form.id == form_id, Form.tenant_id == user["tenant_id"]
    ).first()
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")

    subs = _query_submissions(db, form_id, user["tenant_id"], date_from, date_to, status)
    enum_map = _build_enumerator_map(db, user["tenant_id"])
    rows = _build_rows(subs, enum_map)

    if not rows:
        # Return an empty .dta with no observations
        df = pd.DataFrame()
    else:
        df = pd.DataFrame(rows)

    # Sanitize column names for Stata compatibility
    col_renames = {}
    used_names: set[str] = set()
    for col in df.columns:
        clean = _sanitize_stata_colname(col)
        # Deduplicate
        base = clean
        counter = 2
        while clean in used_names:
            suffix = str(counter)
            clean = base[: 32 - len(suffix)] + suffix
            counter += 1
        used_names.add(clean)
        col_renames[col] = clean
    df.rename(columns=col_renames, inplace=True)

    # Convert object columns to string to avoid Stata type issues
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].astype(str).replace("None", "")

    # Numeric coercion: try to convert string columns that look numeric
    for col in df.columns:
        if df[col].dtype == object:
            try:
                converted = pd.to_numeric(df[col], errors="coerce")
                # Only convert if most values parsed successfully
                if converted.notna().sum() > 0 and converted.isna().sum() <= df[col].eq("").sum():
                    df[col] = converted
            except (ValueError, TypeError):
                pass

    # Write .dta to bytes buffer
    buf = io.BytesIO()
    try:
        df.to_stata(
            buf,
            write_index=False,
            version=118,  # Stata 14+ format — wide compatibility
            convert_dates=None,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate .dta file: {e}")

    buf.seek(0)
    filename = f"{form.title}.dta"

    return StreamingResponse(
        buf,
        media_type="application/x-stata-dta",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Stata alias + SPSS export
# ---------------------------------------------------------------------------

@router.get("/{form_id}/stata")
def export_stata(
    form_id: str,
    date_from: Optional[datetime] = Query(None),
    date_to: Optional[datetime] = Query(None),
    status: Optional[str] = Query(None),
    user: dict = Depends(require_supervisor),
    db: Session = Depends(get_db),
):
    """Export submissions as Stata .dta (alias for /dta endpoint)."""
    return export_dta(form_id=form_id, date_from=date_from, date_to=date_to, status=status, user=user, db=db)


@router.get("/{form_id}/spss")
def export_spss(
    form_id: str,
    date_from: Optional[datetime] = Query(None),
    date_to: Optional[datetime] = Query(None),
    status: Optional[str] = Query(None),
    user: dict = Depends(require_supervisor),
    db: Session = Depends(get_db),
):
    """Export submissions as SPSS .sav file using pyreadstat."""
    try:
        import pyreadstat
    except ImportError:
        raise HTTPException(status_code=501, detail="pyreadstat not installed — run: pip install pyreadstat")

    if not HAS_PANDAS:
        raise HTTPException(status_code=501, detail="SPSS export requires pandas.")

    form = db.query(Form).filter(
        Form.id == form_id, Form.tenant_id == user["tenant_id"]
    ).first()
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")

    subs = _query_submissions(db, form_id, user["tenant_id"], date_from, date_to, status)
    enum_map = _build_enumerator_map(db, user["tenant_id"])
    rows = _build_rows(subs, enum_map)

    df = pd.DataFrame(rows) if rows else pd.DataFrame()

    schema = form.json_schema or {}
    raw_fields = schema.get("fields", [])
    if not raw_fields and schema.get("sections"):
        for sec in schema["sections"]:
            raw_fields.extend(sec.get("fields", []))
    variable_labels = {f.get("id", ""): f.get("label", f.get("id", "")) for f in raw_fields if f.get("id")}

    col_renames: dict = {}
    used_names: set = set()
    for col in df.columns:
        clean = _sanitize_stata_colname(col)
        base = clean
        counter = 2
        while clean in used_names:
            suffix = str(counter)
            clean = base[: 32 - len(suffix)] + suffix
            counter += 1
        used_names.add(clean)
        col_renames[col] = clean
    df.rename(columns=col_renames, inplace=True)

    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].astype(str).replace("None", "")

    import tempfile
    import os
    with tempfile.NamedTemporaryFile(suffix=".sav", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        col_labels = {col_renames.get(k, k): v for k, v in variable_labels.items() if col_renames.get(k, k) in df.columns}
        pyreadstat.write_sav(df, tmp_path, column_labels=col_labels)
        with open(tmp_path, "rb") as f:
            content = f.read()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate .sav file: {e}")
    finally:
        os.unlink(tmp_path)

    filename = f"{form.title}.sav"
    return StreamingResponse(
        iter([content]),
        media_type="application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Google Sheets — connection status check
# ---------------------------------------------------------------------------

@router.get("/sheets/status")
def sheets_status(user: dict = Depends(require_supervisor)):
    """Return whether Google Drive/Sheets OAuth is configured for this server."""
    import os
    from app.core.config import settings
    token_path = settings.GDRIVE_TOKEN_PATH
    configured = os.path.exists(token_path)
    return {"configured": configured}


# ---------------------------------------------------------------------------
# Google Sheets export
# ---------------------------------------------------------------------------

@router.post("/{form_id}/sheets")
def export_sheets(
    form_id: str,
    date_from: Optional[datetime] = Query(None, description="Filter: local_created_at >="),
    date_to: Optional[datetime] = Query(None, description="Filter: local_created_at <="),
    status: Optional[str] = Query(None, description="Filter by submission status"),
    user: dict = Depends(require_supervisor),
    db: Session = Depends(get_db),
):
    """Export submissions to a new Google Sheet and return its URL.

    Requires Google Drive + Sheets OAuth to be configured
    (run scripts/gdrive_auth.py with the updated scopes).
    """
    try:
        from app.services.sheets import export_to_sheets
    except ImportError as e:
        from fastapi import HTTPException as _HTTPException
        raise _HTTPException(status_code=501, detail=f"Google Sheets service unavailable: {e}")

    from fastapi import HTTPException

    form = db.query(Form).filter(
        Form.id == form_id, Form.tenant_id == user["tenant_id"]
    ).first()
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")

    subs = _query_submissions(db, form_id, user["tenant_id"], date_from, date_to, status)
    enum_map = _build_enumerator_map(db, user["tenant_id"])
    rows_dicts = _build_rows(subs, enum_map)

    if not rows_dicts:
        raise HTTPException(status_code=422, detail="No submissions match the selected filters")

    # Build headers (union across all rows)
    all_keys: list[str] = []
    seen: set[str] = set()
    for row in rows_dicts:
        for k in row:
            if k not in seen:
                all_keys.append(k)
                seen.add(k)

    # Convert to list-of-lists for the Sheets API
    value_rows = [[str(row.get(k, "")) for k in all_keys] for row in rows_dicts]

    date_str = datetime.utcnow().strftime("%Y-%m-%d")
    sheet_title = f"{form.title} — {date_str}"

    try:
        url = export_to_sheets(
            title=sheet_title,
            headers=all_keys,
            rows=value_rows,
            tenant_id=str(user["tenant_id"]),
        )
    except RuntimeError as e:
        raise HTTPException(status_code=503, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Sheets export failed: {e}")

    return {"url": url, "rows": len(value_rows), "title": sheet_title}


# ---------------------------------------------------------------------------
# Excel (.xlsx) export
# ---------------------------------------------------------------------------

@router.get("/{form_id}/xlsx")
def export_xlsx(
    form_id: str,
    date_from: Optional[datetime] = Query(None),
    date_to: Optional[datetime] = Query(None),
    status: Optional[str] = Query(None),
    user: dict = Depends(require_supervisor),
    db: Session = Depends(get_db),
):
    """Export submissions as a formatted Excel .xlsx workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        from fastapi import HTTPException as _E
        raise _E(status_code=501, detail="openpyxl not installed — run: pip install openpyxl")

    from fastapi import HTTPException

    form = db.query(Form).filter(
        Form.id == form_id, Form.tenant_id == user["tenant_id"]
    ).first()
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")

    subs = _query_submissions(db, form_id, user["tenant_id"], date_from, date_to, status)
    enum_map = _build_enumerator_map(db, user["tenant_id"])
    rows_dicts = _build_rows(subs, enum_map)

    # Build unified column list
    all_keys: list[str] = []
    seen: set[str] = set()
    for row in rows_dicts:
        for k in row:
            if k not in seen:
                all_keys.append(k)
                seen.add(k)

    # ── Create workbook ──────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Submissions"

    # Header style
    header_fill = PatternFill(start_color="1E1E2E", end_color="1E1E2E", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="CBA6F7", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="313244")
    header_border = Border(bottom=Side(style="medium", color="89B4FA"))

    # Alternate row fills
    fill_odd  = PatternFill(start_color="1E1E2E", end_color="1E1E2E", fill_type="solid")
    fill_even = PatternFill(start_color="181825", end_color="181825", fill_type="solid")
    cell_font = Font(name="Calibri", color="CDD6F4", size=10)

    # Write headers
    for col_idx, key in enumerate(all_keys, start=1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Write data rows
    for row_idx, row_dict in enumerate(rows_dicts, start=2):
        fill = fill_odd if row_idx % 2 == 0 else fill_even
        for col_idx, key in enumerate(all_keys, start=1):
            val = row_dict.get(key, "")
            # Coerce None to empty string for display
            if val is None:
                val = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if not isinstance(val, (int, float)) else val)
            cell.font = cell_font
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")

    # Auto-size columns (cap at 50 chars wide)
    for col_idx, key in enumerate(all_keys, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(len(str(key)), *(len(str(r.get(key, "") or "")) for r in rows_dicts) if rows_dicts else [0])
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # ── Stream response ──────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_title = "".join(c for c in form.title if c.isalnum() or c in " _-")
    filename = f"{safe_title}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Consent report (DPDP compliance)
# ---------------------------------------------------------------------------

@router.get("/consent-report")
def export_consent_report(
    form_id: str = _Query(..., description="Form ID to export consent data for"),
    user: dict = Depends(require_supervisor),
    db: Session = Depends(get_db),
):
    """Export a CSV of consent records for a form (DPDP compliance)."""
    form = db.query(Form).filter(Form.id == form_id, Form.tenant_id == user["tenant_id"]).first()
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")

    subs = (
        db.query(Submission, User.phone)
        .outerjoin(User, Submission.enumerator_id == User.id)
        .filter(Submission.form_id == form_id, Submission.tenant_id == user["tenant_id"])
        .all()
    )

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["submission_id", "enumerator_phone", "submitted_at", "consent_given", "consent_timestamp"])
    for s, phone in subs:
        writer.writerow([
            str(s.id),
            phone or "",
            s.server_received_at.isoformat() if s.server_received_at else "",
            str(s.consent_given) if s.consent_given is not None else "",
            s.consent_timestamp.isoformat() if s.consent_timestamp else "",
        ])
    buf.seek(0)

    safe_title = "".join(c for c in form.title if c.isalnum() or c in " _-")
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="consent_{safe_title}.csv"'},
    )
