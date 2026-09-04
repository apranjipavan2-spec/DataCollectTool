"""
FieldGovern json_schema → XLSForm (.xlsx) serializer.

The inverse of xlsform_parser. Emits the exact survey/choices/settings columns
the parser reads, so a form built or edited in FieldGovern downloads as a clean
XLSForm that re-imports into FieldGovern (or any ODK/Kobo tool) without warnings.

Lossy by nature — a schema no longer carries the original file's AUDIT sheet,
per-list names, or integer constraints. What round-trips: fields, types, labels
(+ Hindi), choices & cascading attrs, choice_filter, relevant (skip logic),
calculate, notes, groups/sections and repeats.
"""
from __future__ import annotations

import io
from typing import Any

import openpyxl

# FG type → XLSForm base type (inverse of parser._TYPE_MAP, picking the
# canonical XLSForm token for each).
_TYPE_OUT: dict[str, str] = {
    "text": "text",
    "number": "integer",
    "decimal": "decimal",
    "date": "date",
    "time": "time",
    "gps": "geopoint",
    "photo": "image",
    "audio": "audio",
    "barcode": "barcode",
    "calculated": "calculate",
    "note": "note",
    "rating": "range",
}

_SURVEY_COLS = ["type", "name", "label", "hint", "required",
                "relevant", "calculation", "appearance", "parameters", "choice_filter"]

# ── skip logic → XLSForm relevant expression ──────────────────────────────────

def _cond_to_xpath(cond: dict) -> str | None:
    """Serialize one condition dict back to an XLSForm relevant fragment."""
    field = cond.get("field")
    op = cond.get("operator")
    val = cond.get("value")
    if not field or not op:
        return None
    ref = f"${{{field}}}"
    if op == "eq":            return f"{ref}='{val}'"
    if op == "neq":           return f"{ref}!='{val}'"
    if op == "is_not_empty":  return f"{ref}!=''"
    if op == "is_empty":      return f"{ref}=''"
    if op == "contains":      return f"selected({ref},'{val}')"
    if op in ("gt", "gte", "lt", "lte"):
        sym = {"gt": ">", "gte": ">=", "lt": "<", "lte": "<="}[op]
        return f"{ref}{sym}{_num(val)}"
    if op in ("age_gte", "age_lt"):
        years = str(val).split("|")[0] if val is not None else "0"
        sym = ">=" if op == "age_gte" else "<"
        return f"age({ref}){sym}{years}"
    return None  # unknown operator — drop rather than emit something un-reimportable


def _num(v: Any) -> str:
    """Render a numeric threshold without a trailing .0 (15.0 -> 15)."""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _group_to_xpath(node: dict) -> str | None:
    """Serialize a skipLogic group (or bare condition) to a relevant expression."""
    if "conditions" not in node:
        return _cond_to_xpath(node)
    logic = node.get("logic", "AND")
    sep = " or " if logic == "OR" else " and "
    parts = []
    for child in node["conditions"]:
        frag = _group_to_xpath(child)
        if not frag:
            continue
        # Parenthesize a nested group of the *other* logic to preserve precedence.
        if "conditions" in child and child.get("logic", "AND") != logic:
            frag = f"({frag})"
        parts.append(frag)
    return sep.join(parts) if parts else None


def _relevant_for(obj: dict) -> str:
    """Prefer the verbatim original (skipLogicRaw); else serialize skipLogic."""
    if obj.get("skipLogicRaw"):
        return str(obj["skipLogicRaw"])
    sl = obj.get("skipLogic")
    return _group_to_xpath(sl) or "" if sl else ""


# ── choice_filter → XLSForm expression ────────────────────────────────────────

def _choice_filter_expr(cf: list[dict]) -> str:
    return " and ".join(f"{c['attr']}=${{{c['field']}}}" for c in cf if c.get("attr") and c.get("field"))


# ── main ──────────────────────────────────────────────────────────────────────

def serialize_xlsform(form_title: str, json_schema: dict, form_id: str = "", version: int = 1) -> bytes:
    """Build XLSForm .xlsx bytes from a FieldGovern json_schema."""
    wb = openpyxl.Workbook()

    survey = wb.active
    survey.title = "survey"
    choices = wb.create_sheet("choices")
    settings = wb.create_sheet("settings")

    # choices accumulate as we walk fields; keep them in encounter order.
    choice_rows: list[dict] = []            # {list_name, name, label, label_hi?, <attr>...}
    choice_attr_cols: list[str] = []        # extra cascading columns (e.g. block, gp)
    any_hindi_choice = False

    survey_rows: list[dict] = []            # dicts keyed by _SURVEY_COLS + optional label_hi
    any_hindi_survey = False

    def emit(**cells: Any) -> None:
        survey_rows.append(cells)

    def walk_field(f: dict) -> None:
        nonlocal any_hindi_survey, any_hindi_choice
        ftype = f.get("type", "text")
        name = f.get("name", "")
        label = f.get("label", name)
        row: dict[str, Any] = {"name": name, "label": label}

        if f.get("label_hi"):
            row["label_hi"] = f["label_hi"]
            any_hindi_survey = True
        if f.get("hint") and not (ftype == "note" and f.get("hint") == label):
            row["hint"] = f["hint"]
        if f.get("required"):
            row["required"] = "yes"
        rel = _relevant_for(f)
        if rel:
            row["relevant"] = rel

        if ftype == "repeat_group":
            emit(type="begin_repeat", **row)
            for child in f.get("fields", []):
                walk_field(child)
            emit(type="end_repeat", name=name)
            return

        if ftype in ("single_choice", "multiple_choice"):
            list_name = name or f"list_{len(choice_rows)}"
            for opt in f.get("options", []):
                crow: dict[str, Any] = {
                    "list_name": list_name,
                    "name": opt.get("value", ""),
                    "label": opt.get("label", opt.get("value", "")),
                }
                if opt.get("label_hi"):
                    crow["label_hi"] = opt["label_hi"]
                    any_hindi_choice = True
                for k, v in opt.items():
                    if k in ("value", "label", "label_hi") or not v:
                        continue
                    if k not in choice_attr_cols:
                        choice_attr_cols.append(k)
                    crow[k] = v
                choice_rows.append(crow)
            kw = "select_one" if ftype == "single_choice" else "select_multiple"
            row["type"] = f"{kw} {list_name}"
            if f.get("choiceFilter"):
                row["choice_filter"] = _choice_filter_expr(f["choiceFilter"])
            emit(**row)
            return

        if ftype == "calculated":
            row["type"] = "calculate"
            if f.get("formula"):
                row["calculation"] = f["formula"]
            emit(**row)
            return

        if ftype == "rating":
            row["type"] = "range"
            row["appearance"] = "rating"
            mn = _num(f["min"]) if f.get("min") is not None else "1"
            mx = _num(f["max"]) if f.get("max") is not None else "5"
            row["parameters"] = f"start={mn} end={mx} step=1"
            emit(**row)
            return

        row["type"] = _TYPE_OUT.get(ftype, "text")
        emit(**row)

    for sec in json_schema.get("sections", []):
        title = sec.get("title", "Section")
        gname = "grp_" + "".join(c if c.isalnum() else "_" for c in title.lower())[:40]
        grow = {"type": "begin_group", "name": gname, "label": title}
        srel = _relevant_for(sec)
        if srel:
            grow["relevant"] = srel
        emit(**grow)
        for f in sec.get("fields", []):
            walk_field(f)
        emit(type="end_group", name=gname)

    # ── write survey sheet ──────────────────────────────────────────────────
    survey_cols = list(_SURVEY_COLS)
    if any_hindi_survey:
        survey_cols.insert(survey_cols.index("label") + 1, "label::Hindi (hi)")
    survey.append(survey_cols)
    for r in survey_rows:
        if "label_hi" in r:
            r = {**r, "label::Hindi (hi)": r.pop("label_hi")}
        survey.append([r.get(c, "") for c in survey_cols])

    # ── write choices sheet ─────────────────────────────────────────────────
    choice_cols = ["list_name", "name", "label"]
    if any_hindi_choice:
        choice_cols.append("label::Hindi (hi)")
    choice_cols += choice_attr_cols
    choices.append(choice_cols)
    for cr in choice_rows:
        if "label_hi" in cr:
            cr = {**cr, "label::Hindi (hi)": cr.pop("label_hi")}
        choices.append([cr.get(c, "") for c in choice_cols])

    # ── write settings sheet ────────────────────────────────────────────────
    settings.append(["form_title", "form_id", "version"])
    settings.append([form_title, form_id or "", version])

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


# ── self-check: schema → xlsx → parse round-trips ─────────────────────────────
if __name__ == "__main__":
    from .xlsform_parser import parse_xlsform

    schema = {
        "title": "Round Trip", "version": 1,
        "sections": [{
            "id": "s1", "title": "Demographics",
            "fields": [
                {"id": "1", "type": "text", "name": "q_name", "label": "Name", "required": True},
                {"id": "2", "type": "number", "name": "q_age", "label": "Age", "required": False},
                {"id": "3", "type": "single_choice", "name": "q_sex", "label": "Sex",
                 "required": True, "options": [
                     {"value": "1", "label": "Male"}, {"value": "2", "label": "Female"}]},
                {"id": "4", "type": "text", "name": "q_spouse", "label": "Spouse name",
                 "skipLogicRaw": "${q_sex}='2'"},
                {"id": "5", "type": "single_choice", "name": "q_gp", "label": "GP",
                 "options": [{"value": "g1", "label": "GP1", "block": "b1"}],
                 "choiceFilter": [{"attr": "block", "field": "q_block"}]},
            ],
        }],
    }
    xlsx = serialize_xlsform(schema["title"], schema, version=1)
    out = parse_xlsform(xlsx, "roundtrip.xlsx")

    names = [f["name"] for s in out["json_schema"]["sections"] for f in s["fields"]]
    assert names == ["q_name", "q_age", "q_sex", "q_spouse", "q_gp"], names
    fields = {f["name"]: f for s in out["json_schema"]["sections"] for f in s["fields"]}
    assert fields["q_sex"]["type"] == "single_choice", fields["q_sex"]
    assert fields["q_sex"]["options"][0] == {"value": "1", "label": "Male"}, fields["q_sex"]["options"]
    assert fields["q_spouse"].get("skipLogic"), "relevant did not round-trip"
    assert fields["q_gp"].get("choiceFilter") == [{"attr": "block", "field": "q_block"}], fields["q_gp"]
    assert fields["q_gp"]["options"][0].get("block") == "b1", "cascading attr lost"
    assert out["json_schema"]["sections"][0]["title"] == "Demographics", "section title lost"
    print("xlsform_serializer self-check: round-trip OK —", len(names), "fields")
