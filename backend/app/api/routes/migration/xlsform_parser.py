"""
XLSForm (.xlsx/.xls) → FieldGovern json_schema parser.

XLSForm standard: https://xlsform.org/en/
Handles: text, integer, decimal, date, time, select_one/multiple,
         geopoint, image, audio, video, barcode, calculate, note,
         range, begin_group/end_group, begin_repeat/end_repeat.
"""
from __future__ import annotations

import io
import re
import uuid
from typing import Any

import openpyxl

from .relevant_parser import parse_relevant

# ── type mapping ─────────────────────────────────────────────────────────────

_TYPE_MAP: dict[str, str] = {
    "text": "text", "string": "text",
    "integer": "number", "int": "number",
    "decimal": "decimal",
    "date": "date",
    "time": "time",
    "datetime": "date",
    "geopoint": "gps", "geotrace": "gps", "geoshape": "gps",
    "image": "photo", "photo": "photo",
    "audio": "audio",
    "video": "audio",
    "barcode": "barcode", "qrcode": "barcode",
    "calculate": "calculated",
    "note": "note",
    "range": "number",
    "select_one": "single_choice",
    "select_multiple": "multiple_choice",
}

_STRUCTURAL = {"begin_group", "end_group", "begin_repeat", "end_repeat"}

# types that warn the user data may be lost or mapped approximately
_WARN_MAP = {
    "video": "video fields mapped to audio — video files not supported",
    "geotrace": "geotrace mapped to gps — only the first point is captured",
    "geoshape": "geoshape mapped to gps — only the first point is captured",
    "datetime": "datetime mapped to date — time component dropped",
}


# ── helpers ──────────────────────────────────────────────────────────────────

def _str(v: Any) -> str:
    return str(v).strip() if v is not None else ""


def _bool(v: Any) -> bool:
    return str(v).strip().lower() in {"yes", "true", "1"}


def _uid() -> str:
    return str(uuid.uuid4())


def _sheet(wb: openpyxl.Workbook, *names: str):
    """Return first sheet whose title matches one of the given names (case-insensitive)."""
    for name in names:
        for title in wb.sheetnames:
            if title.strip().lower() == name.lower():
                return wb[title]
    return None


def _rows_as_dicts(ws) -> list[dict[str, str]]:
    """Convert worksheet to list of dicts using first row as header."""
    if ws is None:
        return []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    # Find the first non-empty row as header
    header_idx = 0
    for i, row in enumerate(rows):
        if any(c is not None for c in row):
            header_idx = i
            break
    headers = [_str(c).lower() for c in rows[header_idx]]
    result = []
    for row in rows[header_idx + 1:]:
        d = {headers[i]: _str(v) for i, v in enumerate(row) if i < len(headers)}
        if any(d.values()):
            result.append(d)
    return result


def _best_label(row: dict, lang_hint: str = "") -> str:
    """Pick the best label column value. Prefers English, falls back to plain label."""
    candidates = [
        f"label::english (en)", f"label::english", f"label::{lang_hint}",
        "label::english (en)", "label",
    ]
    for key in candidates:
        if row.get(key):
            return row[key]
    # fallback: any key starting with "label"
    for k, v in row.items():
        if k.startswith("label") and v:
            return v
    return row.get("name", "")


def _best_hint(row: dict) -> str:
    for k, v in row.items():
        if k.startswith("hint") and v:
            return v
    return ""


def _parse_range_params(row: dict) -> tuple[float | None, float | None]:
    """Extract min/max from XLSForm range parameters field."""
    params = row.get("parameters", "")
    mn = mx = None
    for part in params.split():
        if part.startswith("start="):
            try: mn = float(part[6:])
            except ValueError: pass
        elif part.startswith("end="):
            try: mx = float(part[4:])
            except ValueError: pass
    return mn, mx


# ── choices builder ───────────────────────────────────────────────────────────

# choices columns that are never cascading attributes
_CHOICE_RESERVED = {"list_name", "list name", "name", "label", "image", "media::image"}


def _build_choices(choices_rows: list[dict]) -> dict[str, list[dict]]:
    """Return {list_name: [{value, label, <attr>...}, ...]}.

    Any extra column beyond list_name/name/label (and not a label:: variant) is kept
    on each option as a cascading attribute, so a later field's choice_filter can match
    on it (e.g. a 'block' column on the GP list)."""
    result: dict[str, list[dict]] = {}
    for row in choices_rows:
        list_name = row.get("list_name") or row.get("list name", "")
        name = row.get("name", "")
        label = _best_label(row)
        if not list_name or not name:
            continue
        opt = {"value": name, "label": label or name}
        for k, v in row.items():
            if k in _CHOICE_RESERVED or k.startswith("label") or not v:
                continue
            opt[k] = v
        result.setdefault(list_name, []).append(opt)
    return result


# choice_filter like "block=${block} and gp=${gp_name}" -> [{attr, field}, ...]
_CF_PART = re.compile(r"([A-Za-z0-9_]+)\s*=\s*\$\{([^}]+)\}")


def _parse_choice_filter(expr: str) -> list[dict]:
    return [{"attr": m[1], "field": m[2]} for m in _CF_PART.finditer(expr or "")]


# ── main parser ───────────────────────────────────────────────────────────────

def parse_xlsform(file_bytes: bytes, filename: str = "") -> dict:
    """
    Parse XLSForm Excel bytes → FieldGovern json_schema + warnings list.

    Returns:
    {
      "title": str,
      "json_schema": { "title": str, "sections": [...], "version": 1 },
      "warnings": [str, ...]          # non-fatal issues
    }
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # ── settings ──────────────────────────────────────────────────────────────
    settings_ws = _sheet(wb, "settings")
    settings_rows = _rows_as_dicts(settings_ws)
    settings = settings_rows[0] if settings_rows else {}
    form_title = (
        settings.get("form_title")
        or settings.get("title")
        or filename.rsplit(".", 1)[0]
        or "Imported Form"
    )

    # ── choices ───────────────────────────────────────────────────────────────
    choices_ws = _sheet(wb, "choices")
    choices_map = _build_choices(_rows_as_dicts(choices_ws))

    # ── survey ────────────────────────────────────────────────────────────────
    survey_ws = _sheet(wb, "survey")
    if survey_ws is None:
        raise ValueError("No 'survey' sheet found — this does not appear to be an XLSForm file.")
    survey_rows = _rows_as_dicts(survey_ws)

    warnings: list[str] = []
    sections: list[dict] = []
    current_section: dict | None = None
    repeat_stack: list[dict] = []  # stack for nested repeat_groups

    def _current_fields() -> list:
        """Return the fields list of the innermost active container."""
        if repeat_stack:
            return repeat_stack[-1]["fields"]
        if current_section is not None:
            return current_section["fields"]
        # No section yet — create a default one
        return _ensure_section("General")

    def _ensure_section(title: str = "Section") -> list:
        nonlocal current_section
        if current_section is None or (title != "Section" and current_section["title"] == "Section"):
            current_section = {"id": _uid(), "title": title, "fields": []}
            sections.append(current_section)
        return current_section["fields"]

    for row in survey_rows:
        raw_type = row.get("type", "").strip()
        if not raw_type:
            continue

        name = row.get("name", "").strip()
        label = _best_label(row)
        hint = _best_hint(row)
        required = _bool(row.get("required", ""))
        relevant = row.get("relevant", "")
        calculation = row.get("calculation", "")
        appearance = row.get("appearance", "").lower()
        choice_filter = row.get("choice_filter") or row.get("choice filter", "")

        # Parse type token (may be "select_one list_name")
        type_parts = raw_type.split()
        base_type = type_parts[0].lower()
        list_name = type_parts[1] if len(type_parts) > 1 else ""

        # ── structural rows ──────────────────────────────────────────────────
        if base_type == "begin_group":
            sec_title = label or name or "Group"
            current_section = {"id": _uid(), "title": sec_title, "fields": []}
            # section-level skip logic (relevant on the begin_group row) — the
            # renderer hides the whole section when this is false.
            if relevant:
                sec_skip, sec_raw = parse_relevant(relevant)
                if sec_skip:
                    current_section["skipLogic"] = sec_skip
                if sec_raw:
                    warnings.append(
                        f"Section '{sec_title}': relevant expression could not be fully "
                        f"parsed — stored raw: '{sec_raw}'"
                    )
            sections.append(current_section)
            continue

        if base_type == "end_group":
            # Close current group; next fields go to a new implicit section
            current_section = None
            continue

        if base_type == "begin_repeat":
            repeat_field: dict = {
                "id": _uid(),
                "type": "repeat_group",
                "name": name,
                "label": label or name,
                "hint": hint,
                "required": required,
                "fields": [],
            }
            _ensure_section()
            _current_fields().append(repeat_field)
            repeat_stack.append(repeat_field)
            continue

        if base_type == "end_repeat":
            if repeat_stack:
                repeat_stack.pop()
            continue

        # ── unsupported / skip ───────────────────────────────────────────────
        if base_type in {"xml-external", "trigger", "hidden", "file"}:
            warnings.append(f"Field '{name}' (type: {raw_type}) skipped — not supported.")
            continue

        # ── map type ─────────────────────────────────────────────────────────
        fg_type = _TYPE_MAP.get(base_type, "text")
        if base_type not in _TYPE_MAP and base_type not in {"select_one", "select_multiple"}:
            warnings.append(f"Field '{name}' has unknown type '{raw_type}' — mapped to text.")

        # warn for lossy mappings
        if base_type in _WARN_MAP:
            warnings.append(f"Field '{name}': {_WARN_MAP[base_type]}")

        # range with rating appearance → rating type
        if base_type == "range" and "rating" in appearance:
            fg_type = "rating"

        # ── build field ───────────────────────────────────────────────────────
        field: dict[str, Any] = {
            "id": _uid(),
            "type": fg_type,
            "name": name,
            "label": label or name,
            "required": required,
        }

        if hint:
            field["hint"] = hint

        # options (select_one / select_multiple)
        if fg_type in {"single_choice", "multiple_choice"}:
            opts = choices_map.get(list_name, [])
            field["options"] = opts
            if not opts:
                warnings.append(f"Field '{name}': choice list '{list_name}' not found in choices sheet.")
            # cascading select: filter options by a parent field's answer
            if choice_filter:
                cf = _parse_choice_filter(choice_filter)
                if cf:
                    field["choiceFilter"] = cf
                else:
                    warnings.append(
                        f"Field '{name}': choice_filter '{choice_filter}' not understood — "
                        f"cascading not applied (all options will show)."
                    )

        # calculated formula
        if fg_type == "calculated" and calculation:
            field["formula"] = calculation

        # number min/max (from range parameters or constraint)
        if base_type == "range":
            mn, mx = _parse_range_params(row)
            if mn is not None: field["min"] = mn
            if mx is not None: field["max"] = mx

        # skip logic from relevant
        if relevant:
            skip_logic, raw_warn = parse_relevant(relevant)
            if skip_logic:
                field["skipLogic"] = skip_logic
            if raw_warn:
                warnings.append(
                    f"Field '{name}': relevant expression could not be fully parsed — "
                    f"stored as skipLogicRaw: '{raw_warn}'"
                )
                field["skipLogicRaw"] = relevant

        # note type: label becomes hint (notes are display-only)
        if fg_type == "note":
            field["hint"] = label or hint
            field["label"] = label or name
            field.pop("required", None)

        # ensure a section exists
        _ensure_section()
        _current_fields().append(field)

    wb.close()

    if not sections:
        sections = [{"id": _uid(), "title": "General", "fields": []}]

    json_schema = {
        "title": form_title,
        "sections": sections,
        "version": 1,
    }

    return {
        "title": form_title,
        "json_schema": json_schema,
        "warnings": warnings,
        "field_count": sum(len(s.get("fields", [])) for s in sections),
        "section_count": len(sections),
    }
