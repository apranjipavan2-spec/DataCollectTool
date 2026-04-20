"""Program management: programs, participant types, questionnaires, location targets, progress reporting."""
import io
from datetime import date, datetime, timedelta
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.deps import require_enumerator, require_supervisor
from app.models.form import Form
from app.models.program import (
    Program, ProgramLocation, ProgramParticipantType,
    ProgramQuestionnaire, QuestionnaireLocationTarget,
)
from app.models.submission import Submission

router = APIRouter()


# ── Schemas ─────────────────────────────────────────────────────────────────

class LocationIn(BaseModel):
    state: str = ""
    district: str
    block: str = ""
    village: str = ""
    gps_lat: Optional[float] = None
    gps_lng: Optional[float] = None

class ProgramIn(BaseModel):
    name: str
    scheme_name: str = ""
    description: str = ""
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    status: str = "active"

class ParticipantTypeIn(BaseModel):
    name: str
    description: str = ""
    sort_order: int = 0

class QuestionnaireIn(BaseModel):
    participant_type_id: Optional[str] = None
    form_id: Optional[str] = None
    name: str
    total_target: int = 0
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    status: str = "active"

class LocationTargetIn(BaseModel):
    location_id: str
    target_count: int
    deadline: Optional[date] = None


# ── Helper: status from counts + deadline ────────────────────────────────────

def _target_status(collected: int, target: int, deadline: Optional[date]) -> str:
    if collected >= target and target > 0:
        return "completed"
    today = date.today()
    if deadline and deadline < today:
        return "overdue"
    pct = (collected / target * 100) if target > 0 else 0
    if deadline and (deadline - today).days <= 7 and pct < 70:
        return "at_risk"
    if collected == 0:
        return "not_started"
    return "in_progress"


# ── Locations (master list per tenant) ──────────────────────────────────────

@router.get("/locations")
def list_locations(user=Depends(require_enumerator), db: Session = Depends(get_db)):
    locs = db.query(ProgramLocation).filter(ProgramLocation.tenant_id == user["tenant_id"]).order_by(
        ProgramLocation.state, ProgramLocation.district, ProgramLocation.block, ProgramLocation.village
    ).all()
    return [{"id": str(l.id), "state": l.state, "district": l.district,
             "block": l.block, "village": l.village,
             "gps_lat": l.gps_lat, "gps_lng": l.gps_lng} for l in locs]


@router.post("/locations")
def create_location(body: LocationIn, user=Depends(require_supervisor), db: Session = Depends(get_db)):
    loc = ProgramLocation(tenant_id=user["tenant_id"], **body.model_dump())
    db.add(loc); db.commit(); db.refresh(loc)
    return {"id": str(loc.id)}


@router.patch("/locations/{loc_id}")
def update_location(loc_id: str, body: LocationIn, user=Depends(require_supervisor), db: Session = Depends(get_db)):
    loc = db.query(ProgramLocation).filter(ProgramLocation.id == loc_id, ProgramLocation.tenant_id == user["tenant_id"]).first()
    if not loc: raise HTTPException(404, "Location not found")
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(loc, k, v)
    db.commit()
    return {"id": str(loc.id)}


@router.delete("/locations/{loc_id}")
def delete_location(loc_id: str, user=Depends(require_supervisor), db: Session = Depends(get_db)):
    loc = db.query(ProgramLocation).filter(ProgramLocation.id == loc_id, ProgramLocation.tenant_id == user["tenant_id"]).first()
    if not loc: raise HTTPException(404, "Location not found")
    db.delete(loc); db.commit()
    return {"deleted": True}


# ── Cross-program overview (must be before /{prog_id} routes) ────────────────

@router.get("/overview")
def progress_overview(user=Depends(require_supervisor), db: Session = Depends(get_db)):
    """Cross-program summary for the dashboard."""
    programs = db.query(Program).filter(
        Program.tenant_id == user["tenant_id"],
        Program.status.in_(["active", "planning"]),
    ).all()
    result = []
    for p in programs:
        rows = _build_progress_rows(str(p.id), user["tenant_id"], db)
        tt = sum(r["target"] for r in rows)
        tc = sum(r["collected"] for r in rows)
        result.append({
            "id": str(p.id), "name": p.name, "scheme_name": p.scheme_name,
            "status": p.status, "total_target": tt, "total_collected": tc,
            "pct": round(tc / tt * 100) if tt else 0,
            "overdue": sum(1 for r in rows if r["status"] == "overdue"),
            "at_risk": sum(1 for r in rows if r["status"] == "at_risk"),
        })
    return result


# ── Programs CRUD ────────────────────────────────────────────────────────────

@router.get("/")
def list_programs(
    scheme: Optional[str] = None,
    status: Optional[str] = None,
    user=Depends(require_enumerator), db: Session = Depends(get_db)
):
    q = db.query(Program).filter(Program.tenant_id == user["tenant_id"])
    if scheme: q = q.filter(Program.scheme_name.ilike(f"%{scheme}%"))
    if status: q = q.filter(Program.status == status)
    programs = q.order_by(Program.created_at.desc()).all()

    result = []
    for p in programs:
        # Quick progress aggregate
        questionnaires = db.query(ProgramQuestionnaire).filter(ProgramQuestionnaire.program_id == p.id).all()
        total_target = sum(q2.total_target for q2 in questionnaires)
        total_collected = 0
        for q2 in questionnaires:
            total_collected += db.query(func.count(Submission.id)).filter(
                Submission.questionnaire_id == q2.id,
                Submission.tenant_id == user["tenant_id"],
            ).scalar() or 0
        result.append({
            "id": str(p.id), "name": p.name, "scheme_name": p.scheme_name,
            "description": p.description, "status": p.status,
            "start_date": p.start_date.isoformat() if p.start_date else None,
            "end_date": p.end_date.isoformat() if p.end_date else None,
            "total_target": total_target, "total_collected": total_collected,
            "pct": round(total_collected / total_target * 100) if total_target else 0,
            "created_at": p.created_at.isoformat() if p.created_at else "",
        })
    return result


@router.post("/")
def create_program(body: ProgramIn, user=Depends(require_supervisor), db: Session = Depends(get_db)):
    p = Program(tenant_id=user["tenant_id"], created_by=user["sub"], **body.model_dump())
    db.add(p); db.commit(); db.refresh(p)
    return {"id": str(p.id)}


@router.get("/{prog_id}")
def get_program(prog_id: str, user=Depends(require_enumerator), db: Session = Depends(get_db)):
    p = db.query(Program).filter(Program.id == prog_id, Program.tenant_id == user["tenant_id"]).first()
    if not p: raise HTTPException(404, "Program not found")

    types = db.query(ProgramParticipantType).filter(ProgramParticipantType.program_id == p.id).order_by(ProgramParticipantType.sort_order).all()
    questionnaires = db.query(ProgramQuestionnaire).filter(ProgramQuestionnaire.program_id == p.id).all()
    form_ids = {q2.form_id for q2 in questionnaires if q2.form_id}
    form_map = {str(f.id): f.title for f in db.query(Form.id, Form.title).filter(Form.id.in_(form_ids)).all()} if form_ids else {}
    type_map = {str(t.id): t.name for t in types}

    qs_data = []
    for q2 in questionnaires:
        targets = db.query(QuestionnaireLocationTarget).filter(QuestionnaireLocationTarget.questionnaire_id == q2.id).all()
        qs_data.append({
            "id": str(q2.id), "name": q2.name,
            "participant_type_id": str(q2.participant_type_id) if q2.participant_type_id else None,
            "participant_type_name": type_map.get(str(q2.participant_type_id), "") if q2.participant_type_id else "",
            "form_id": str(q2.form_id) if q2.form_id else None,
            "form_title": form_map.get(str(q2.form_id), "") if q2.form_id else "",
            "total_target": q2.total_target, "status": q2.status,
            "start_date": q2.start_date.isoformat() if q2.start_date else None,
            "end_date": q2.end_date.isoformat() if q2.end_date else None,
            "location_targets": [
                {"id": str(t.id), "location_id": str(t.location_id),
                 "target_count": t.target_count,
                 "deadline": t.deadline.isoformat() if t.deadline else None}
                for t in targets
            ],
        })

    return {
        "id": str(p.id), "name": p.name, "scheme_name": p.scheme_name,
        "description": p.description, "status": p.status,
        "start_date": p.start_date.isoformat() if p.start_date else None,
        "end_date": p.end_date.isoformat() if p.end_date else None,
        "participant_types": [
            {"id": str(t.id), "name": t.name, "description": t.description, "sort_order": t.sort_order}
            for t in types
        ],
        "questionnaires": qs_data,
    }


@router.patch("/{prog_id}")
def update_program(prog_id: str, body: ProgramIn, user=Depends(require_supervisor), db: Session = Depends(get_db)):
    p = db.query(Program).filter(Program.id == prog_id, Program.tenant_id == user["tenant_id"]).first()
    if not p: raise HTTPException(404, "Program not found")
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(p, k, v)
    db.commit()
    return {"id": str(p.id)}


@router.delete("/{prog_id}")
def delete_program(prog_id: str, user=Depends(require_supervisor), db: Session = Depends(get_db)):
    p = db.query(Program).filter(Program.id == prog_id, Program.tenant_id == user["tenant_id"]).first()
    if not p: raise HTTPException(404, "Program not found")
    db.delete(p); db.commit()
    return {"deleted": True}


# ── Participant Types ─────────────────────────────────────────────────────────

@router.post("/{prog_id}/participant-types")
def create_participant_type(prog_id: str, body: ParticipantTypeIn, user=Depends(require_supervisor), db: Session = Depends(get_db)):
    p = db.query(Program).filter(Program.id == prog_id, Program.tenant_id == user["tenant_id"]).first()
    if not p: raise HTTPException(404, "Program not found")
    pt = ProgramParticipantType(program_id=prog_id, tenant_id=user["tenant_id"], **body.model_dump())
    db.add(pt); db.commit(); db.refresh(pt)
    return {"id": str(pt.id)}


@router.patch("/{prog_id}/participant-types/{pt_id}")
def update_participant_type(prog_id: str, pt_id: str, body: ParticipantTypeIn, user=Depends(require_supervisor), db: Session = Depends(get_db)):
    pt = db.query(ProgramParticipantType).filter(ProgramParticipantType.id == pt_id, ProgramParticipantType.tenant_id == user["tenant_id"]).first()
    if not pt: raise HTTPException(404, "Participant type not found")
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(pt, k, v)
    db.commit()
    return {"id": str(pt.id)}


@router.delete("/{prog_id}/participant-types/{pt_id}")
def delete_participant_type(prog_id: str, pt_id: str, user=Depends(require_supervisor), db: Session = Depends(get_db)):
    pt = db.query(ProgramParticipantType).filter(ProgramParticipantType.id == pt_id, ProgramParticipantType.tenant_id == user["tenant_id"]).first()
    if not pt: raise HTTPException(404, "Participant type not found")
    db.delete(pt); db.commit()
    return {"deleted": True}


# ── Questionnaires ────────────────────────────────────────────────────────────

@router.post("/{prog_id}/questionnaires")
def create_questionnaire(prog_id: str, body: QuestionnaireIn, user=Depends(require_supervisor), db: Session = Depends(get_db)):
    p = db.query(Program).filter(Program.id == prog_id, Program.tenant_id == user["tenant_id"]).first()
    if not p: raise HTTPException(404, "Program not found")
    q = ProgramQuestionnaire(program_id=prog_id, tenant_id=user["tenant_id"], **body.model_dump())
    db.add(q); db.commit(); db.refresh(q)
    return {"id": str(q.id)}


@router.patch("/{prog_id}/questionnaires/{q_id}")
def update_questionnaire(prog_id: str, q_id: str, body: QuestionnaireIn, user=Depends(require_supervisor), db: Session = Depends(get_db)):
    q = db.query(ProgramQuestionnaire).filter(ProgramQuestionnaire.id == q_id, ProgramQuestionnaire.tenant_id == user["tenant_id"]).first()
    if not q: raise HTTPException(404, "Questionnaire not found")
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(q, k, v)
    db.commit()
    return {"id": str(q.id)}


@router.delete("/{prog_id}/questionnaires/{q_id}")
def delete_questionnaire(prog_id: str, q_id: str, user=Depends(require_supervisor), db: Session = Depends(get_db)):
    q = db.query(ProgramQuestionnaire).filter(ProgramQuestionnaire.id == q_id, ProgramQuestionnaire.tenant_id == user["tenant_id"]).first()
    if not q: raise HTTPException(404, "Questionnaire not found")
    db.delete(q); db.commit()
    return {"deleted": True}


# ── Location Targets ──────────────────────────────────────────────────────────

@router.post("/{prog_id}/questionnaires/{q_id}/targets")
def create_location_target(prog_id: str, q_id: str, body: LocationTargetIn, user=Depends(require_supervisor), db: Session = Depends(get_db)):
    q = db.query(ProgramQuestionnaire).filter(ProgramQuestionnaire.id == q_id, ProgramQuestionnaire.tenant_id == user["tenant_id"]).first()
    if not q: raise HTTPException(404, "Questionnaire not found")
    t = QuestionnaireLocationTarget(questionnaire_id=q_id, tenant_id=user["tenant_id"], **body.model_dump())
    db.add(t); db.commit(); db.refresh(t)
    return {"id": str(t.id)}


@router.patch("/{prog_id}/questionnaires/{q_id}/targets/{t_id}")
def update_location_target(prog_id: str, q_id: str, t_id: str, body: LocationTargetIn, user=Depends(require_supervisor), db: Session = Depends(get_db)):
    t = db.query(QuestionnaireLocationTarget).filter(QuestionnaireLocationTarget.id == t_id, QuestionnaireLocationTarget.tenant_id == user["tenant_id"]).first()
    if not t: raise HTTPException(404, "Target not found")
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(t, k, v)
    db.commit()
    return {"id": str(t.id)}


@router.delete("/{prog_id}/questionnaires/{q_id}/targets/{t_id}")
def delete_location_target(prog_id: str, q_id: str, t_id: str, user=Depends(require_supervisor), db: Session = Depends(get_db)):
    t = db.query(QuestionnaireLocationTarget).filter(QuestionnaireLocationTarget.id == t_id, QuestionnaireLocationTarget.tenant_id == user["tenant_id"]).first()
    if not t: raise HTTPException(404, "Target not found")
    db.delete(t); db.commit()
    return {"deleted": True}


# ── Progress ──────────────────────────────────────────────────────────────────

def _build_progress_rows(prog_id: str, tenant_id: str, db: Session,
                          participant_type_id: Optional[str] = None,
                          questionnaire_id: Optional[str] = None,
                          district: Optional[str] = None,
                          block: Optional[str] = None,
                          status_filter: Optional[str] = None) -> list[dict]:
    q_query = db.query(ProgramQuestionnaire).filter(
        ProgramQuestionnaire.program_id == prog_id,
        ProgramQuestionnaire.tenant_id == tenant_id,
    )
    if participant_type_id:
        q_query = q_query.filter(ProgramQuestionnaire.participant_type_id == participant_type_id)
    if questionnaire_id:
        q_query = q_query.filter(ProgramQuestionnaire.id == questionnaire_id)
    questionnaires = q_query.all()

    type_ids = {q2.participant_type_id for q2 in questionnaires if q2.participant_type_id}
    type_map = {}
    if type_ids:
        types = db.query(ProgramParticipantType).filter(ProgramParticipantType.id.in_(type_ids)).all()
        type_map = {str(t.id): t.name for t in types}

    rows = []
    for q2 in questionnaires:
        targets = db.query(QuestionnaireLocationTarget, ProgramLocation).join(
            ProgramLocation, QuestionnaireLocationTarget.location_id == ProgramLocation.id
        ).filter(QuestionnaireLocationTarget.questionnaire_id == q2.id)

        if district:
            targets = targets.filter(ProgramLocation.district.ilike(f"%{district}%"))
        if block:
            targets = targets.filter(ProgramLocation.block.ilike(f"%{block}%"))

        for target, loc in targets.all():
            collected = db.query(func.count(Submission.id)).filter(
                Submission.questionnaire_id == q2.id,
                Submission.location_id == target.location_id,
                Submission.tenant_id == tenant_id,
            ).scalar() or 0

            # Fallback: count by form_id+date range if no tagged submissions
            if collected == 0 and q2.form_id:
                fq = db.query(func.count(Submission.id)).filter(
                    Submission.form_id == q2.form_id,
                    Submission.tenant_id == tenant_id,
                )
                if q2.start_date:
                    fq = fq.filter(Submission.local_created_at >= datetime.combine(q2.start_date, datetime.min.time()))
                if q2.end_date:
                    fq = fq.filter(Submission.local_created_at <= datetime.combine(q2.end_date, datetime.max.time()))
                collected = fq.scalar() or 0

            st = _target_status(collected, target.target_count, target.deadline)
            if status_filter and st != status_filter:
                continue

            days_remaining = None
            if target.deadline:
                days_remaining = (target.deadline - date.today()).days

            rows.append({
                "target_id": str(target.id),
                "questionnaire_id": str(q2.id),
                "questionnaire_name": q2.name,
                "participant_type_id": str(q2.participant_type_id) if q2.participant_type_id else None,
                "participant_type_name": type_map.get(str(q2.participant_type_id), "") if q2.participant_type_id else "",
                "location_id": str(loc.id),
                "state": loc.state, "district": loc.district,
                "block": loc.block, "village": loc.village,
                "target": target.target_count,
                "collected": collected,
                "remaining": max(0, target.target_count - collected),
                "pct": round(collected / target.target_count * 100) if target.target_count else 0,
                "deadline": target.deadline.isoformat() if target.deadline else None,
                "days_remaining": days_remaining,
                "status": st,
            })
    return rows


@router.get("/{prog_id}/progress")
def get_progress(
    prog_id: str,
    participant_type_id: Optional[str] = None,
    questionnaire_id: Optional[str] = None,
    district: Optional[str] = None,
    block: Optional[str] = None,
    status: Optional[str] = None,
    user=Depends(require_enumerator), db: Session = Depends(get_db),
):
    p = db.query(Program).filter(Program.id == prog_id, Program.tenant_id == user["tenant_id"]).first()
    if not p: raise HTTPException(404, "Program not found")

    rows = _build_progress_rows(prog_id, user["tenant_id"], db,
                                participant_type_id, questionnaire_id, district, block, status)
    total_target = sum(r["target"] for r in rows)
    total_collected = sum(r["collected"] for r in rows)
    overdue = sum(1 for r in rows if r["status"] == "overdue")
    at_risk = sum(1 for r in rows if r["status"] == "at_risk")

    return {
        "program_id": prog_id,
        "program_name": p.name,
        "scheme_name": p.scheme_name,
        "summary": {
            "total_target": total_target,
            "total_collected": total_collected,
            "remaining": max(0, total_target - total_collected),
            "pct": round(total_collected / total_target * 100) if total_target else 0,
            "overdue": overdue,
            "at_risk": at_risk,
        },
        "rows": rows,
    }


# ── Excel export ──────────────────────────────────────────────────────────────

@router.get("/{prog_id}/progress/xlsx")
def export_progress_xlsx(
    prog_id: str,
    participant_type_id: Optional[str] = None,
    questionnaire_id: Optional[str] = None,
    district: Optional[str] = None,
    block: Optional[str] = None,
    status: Optional[str] = None,
    user=Depends(require_supervisor), db: Session = Depends(get_db),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(501, "openpyxl not installed")

    p = db.query(Program).filter(Program.id == prog_id, Program.tenant_id == user["tenant_id"]).first()
    if not p: raise HTTPException(404, "Program not found")

    rows = _build_progress_rows(prog_id, user["tenant_id"], db, participant_type_id, questionnaire_id, district, block, status)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Progress Report"

    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center")

    headers = ["Questionnaire", "Participant Type", "State", "District", "Block", "Village",
               "Target", "Collected", "Remaining", "%", "Status", "Deadline", "Days Left"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = center

    STATUS_COLOR = {
        "completed": "D1FAE5", "in_progress": "DBEAFE", "not_started": "F3F4F6",
        "at_risk": "FEF3C7", "overdue": "FEE2E2",
    }
    for ri, row in enumerate(rows, 2):
        vals = [row["questionnaire_name"], row["participant_type_name"], row["state"],
                row["district"], row["block"], row["village"], row["target"],
                row["collected"], row["remaining"], row["pct"], row["status"],
                row["deadline"] or "", row["days_remaining"] if row["days_remaining"] is not None else ""]
        fill = PatternFill(start_color=STATUS_COLOR.get(row["status"], "FFFFFF"),
                           end_color=STATUS_COLOR.get(row["status"], "FFFFFF"), fill_type="solid")
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.fill = fill

    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    tt = sum(r["target"] for r in rows); tc = sum(r["collected"] for r in rows)
    for row in [["Program", p.name], ["Scheme", p.scheme_name], ["Generated", date.today().isoformat()],
                ["Total Target", tt], ["Collected", tc], ["Remaining", tt - tc],
                ["% Complete", f"{round(tc/tt*100) if tt else 0}%"],
                ["Overdue Targets", sum(1 for r in rows if r["status"] == "overdue")]]:
        ws2.append(row)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    safe = "".join(c for c in p.name if c.isalnum() or c in " _-")
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{safe}_progress.xlsx"'})


# ── PDF export ────────────────────────────────────────────────────────────────

@router.get("/{prog_id}/progress/pdf")
def export_progress_pdf(
    prog_id: str,
    participant_type_id: Optional[str] = None,
    questionnaire_id: Optional[str] = None,
    district: Optional[str] = None,
    block: Optional[str] = None,
    status: Optional[str] = None,
    user=Depends(require_supervisor), db: Session = Depends(get_db),
):
    p = db.query(Program).filter(Program.id == prog_id, Program.tenant_id == user["tenant_id"]).first()
    if not p: raise HTTPException(404, "Program not found")

    rows = _build_progress_rows(prog_id, user["tenant_id"], db, participant_type_id, questionnaire_id, district, block, status)
    tt = sum(r["target"] for r in rows); tc = sum(r["collected"] for r in rows)
    pct = round(tc / tt * 100) if tt else 0

    STATUS_COLOR = {"completed": "#D1FAE5", "in_progress": "#DBEAFE", "not_started": "#F3F4F6",
                    "at_risk": "#FEF3C7", "overdue": "#FEE2E2"}

    row_html = "".join(
        f"<tr style='background:{STATUS_COLOR.get(r[\"status\"],\"#fff\")}'>"
        f"<td>{r['questionnaire_name']}</td><td>{r['participant_type_name']}</td>"
        f"<td>{r['district']}</td><td>{r['block']}</td><td>{r['village']}</td>"
        f"<td style='text-align:center'>{r['target']}</td>"
        f"<td style='text-align:center'>{r['collected']}</td>"
        f"<td style='text-align:center'>{r['remaining']}</td>"
        f"<td style='text-align:center'>{r['pct']}%</td>"
        f"<td style='text-align:center'><span style='padding:2px 8px;border-radius:4px;font-size:11px;"
        f"background:{STATUS_COLOR.get(r[\"status\"],\"#eee\")}'>{r['status'].replace('_',' ').title()}</span></td>"
        f"<td style='text-align:center'>{r['deadline'] or '—'}</td>"
        f"</tr>"
        for r in rows
    )

    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<title>Progress Report — {p.name}</title>
<style>
  body{{font-family:Arial,sans-serif;font-size:12px;color:#111;padding:24px}}
  h1{{font-size:20px;margin-bottom:4px}} h2{{font-size:13px;color:#555;font-weight:normal;margin-top:0}}
  .summary{{display:flex;gap:16px;margin:16px 0}}
  .card{{border:1px solid #ddd;border-radius:6px;padding:12px 16px;min-width:100px;text-align:center}}
  .card .num{{font-size:24px;font-weight:bold}} .card .lbl{{font-size:11px;color:#666}}
  table{{width:100%;border-collapse:collapse;margin-top:16px}}
  th{{background:#1E40AF;color:#fff;padding:8px 6px;text-align:left;font-size:11px}}
  td{{padding:6px;border-bottom:1px solid #eee;font-size:11px}}
  @media print{{.no-print{{display:none}}}}
</style></head><body>
<h1>Progress Report: {p.name}</h1>
<h2>Scheme: {p.scheme_name or '—'} &nbsp;|&nbsp; Generated: {date.today().isoformat()}</h2>
<div class="summary">
  <div class="card"><div class="num">{tt}</div><div class="lbl">Total Target</div></div>
  <div class="card"><div class="num">{tc}</div><div class="lbl">Collected</div></div>
  <div class="card"><div class="num">{tt-tc}</div><div class="lbl">Remaining</div></div>
  <div class="card"><div class="num">{pct}%</div><div class="lbl">Complete</div></div>
  <div class="card"><div class="num">{sum(1 for r in rows if r['status']=='overdue')}</div><div class="lbl">Overdue</div></div>
  <div class="card"><div class="num">{sum(1 for r in rows if r['status']=='at_risk')}</div><div class="lbl">At Risk</div></div>
</div>
<button class="no-print" onclick="window.print()" style="padding:8px 16px;background:#1E40AF;color:#fff;border:none;border-radius:4px;cursor:pointer">Print / Save as PDF</button>
<table><tr>
  <th>Questionnaire</th><th>Participant Type</th><th>District</th><th>Block</th><th>Village</th>
  <th>Target</th><th>Collected</th><th>Remaining</th><th>%</th><th>Status</th><th>Deadline</th>
</tr>{row_html}</table>
</body></html>"""

    buf = io.BytesIO(html.encode())
    safe = "".join(c for c in p.name if c.isalnum() or c in " _-")
    return StreamingResponse(buf, media_type="text/html",
                             headers={"Content-Disposition": f'attachment; filename="{safe}_progress.html"'})
