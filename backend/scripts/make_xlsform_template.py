"""Generate the canonical FieldGovern XLSForm starter template.

Produces `FieldGovern_Form_Template.xlsx` with the three sheets our importer
(app/api/routes/migration/xlsform_parser.py) reads: survey, choices, settings.
Every column the parser looks at is present, with one filled-in example of each
supported field type so a non-technical user can copy a row and edit it.

Run:  python -m scripts.make_xlsform_template  [output_path]
"""
import sys

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NOTE_FONT = Font(color="808080", italic=True)


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for c in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 22


def build() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    # ── survey sheet ─────────────────────────────────────────────────────────
    survey = wb.active
    survey.title = "survey"
    survey_cols = ["type", "name", "label", "hint", "required",
                   "relevant", "calculation", "appearance", "parameters"]
    survey.append(survey_cols)
    survey_rows = [
        # type, name, label, hint, required, relevant, calculation, appearance, parameters
        ["text", "respondent_name", "Full name", "As on ID", "yes", "", "", "", ""],
        ["integer", "age", "Age (years)", "", "yes", "", "", "", ""],
        ["decimal", "land_acres", "Land area (acres)", "", "no", "", "", "", ""],
        ["date", "interview_date", "Date of interview", "", "yes", "", "", "", ""],
        ["time", "start_time", "Start time", "", "no", "", "", "", ""],
        ["select_one gender", "gender", "Gender", "", "yes", "", "", "", ""],
        ["select_multiple issues", "issues", "Health issues", "Select all that apply", "no", "", "", "", ""],
        ["geopoint", "location", "GPS location", "", "no", "", "", "", ""],
        ["image", "house_photo", "Photo of house", "", "no", "", "", "", ""],
        ["barcode", "hh_id", "Household barcode", "", "no", "", "", "", ""],
        ["note", "thanks_note", "Thank you for your time.", "", "", "", "", "", ""],
        ["calculate", "age_next_year", "", "", "", "", "age + 1", "", ""],
        ["range", "satisfaction", "Satisfaction (1-5)", "", "no", "", "", "rating", "start=1 end=5 step=1"],
        # relevant = skip logic: only show if gender is female
        ["integer", "num_children", "Number of children", "", "no", "${gender} = 'female'", "", "", ""],
        # a repeat group (roster): one block repeated per household member
        ["begin_repeat", "members", "Household members", "", "", "", "", "", ""],
        ["text", "member_name", "Member name", "", "yes", "", "", "", ""],
        ["integer", "member_age", "Member age", "", "no", "", "", "", ""],
        ["end_repeat", "members", "", "", "", "", "", "", ""],
    ]
    for r in survey_rows:
        survey.append(r)
    _style_header(survey, len(survey_cols))

    # ── choices sheet ────────────────────────────────────────────────────────
    choices = wb.create_sheet("choices")
    choices_cols = ["list_name", "name", "label"]
    choices.append(choices_cols)
    choices_rows = [
        ["gender", "female", "Female"],
        ["gender", "male", "Male"],
        ["gender", "other", "Other"],
        ["issues", "fever", "Fever / Malaria"],
        ["issues", "cough", "Cough / Respiratory"],
        ["issues", "chronic", "Chronic (BP, diabetes)"],
        ["issues", "none", "None"],
    ]
    for r in choices_rows:
        choices.append(r)
    _style_header(choices, len(choices_cols))

    # ── settings sheet ───────────────────────────────────────────────────────
    settings = wb.create_sheet("settings")
    settings_cols = ["form_title", "form_id", "version"]
    settings.append(settings_cols)
    settings.append(["My Survey", "my_survey", "1"])
    _style_header(settings, len(settings_cols))

    return wb


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else "FieldGovern_Form_Template.xlsx"
    build().save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
