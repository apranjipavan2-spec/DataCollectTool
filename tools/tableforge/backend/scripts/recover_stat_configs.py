"""One-off recovery helper (not wired into any route): given a tableforge project
file plus any number of sibling files/version-histories to search as fallback
sources, reports which columns each stat-model table (correlation, crosstab,
t-test, ANOVA, etc.) actually used.

Read-only. Never prints row-level survey data — only table titles, stat
types, and the column NAMES a table was built from (survey question labels).

Usage: python3 recover_stat_configs.py <current.tableforge> [other.tableforge ...] [--source-xlsx <path>]

--source-xlsx reads ONLY the header row of the actual source data file (no
row data) and merges those real column names into the matching pool — this
covers columns that no surviving table happens to reference anymore, which
the project-file-only pool can't see.
"""
import json
import re
import sys
from pathlib import Path

STAT_TITLES = {
    "correlation": "Correlation Matrix (Pearson)",
    "descriptive": "Descriptive Statistics (Table 1)",
    "crosstab": "Cross-Tabulation + Cramér’s V",
    "ttest": "t-Test (Welch) + Cohen’s d",
    "anova": "One-Way ANOVA + η²/ω²",
    "regression": "OLS Regression",
    "normality": "Normality Tests",
    "outlier": "Outlier Detection",
    "frequency": "Frequency Distribution",
    "paired_ttest": "Paired t-Test (Pre vs Post)",
    "wilcoxon": "Wilcoxon Signed-Rank (paired, non-parametric)",
    "mcnemar": "McNemar’s Test (paired binary)",
    "kruskal": "Kruskal-Wallis (non-parametric ANOVA)",
    "friedman": "Friedman Test (repeated measures)",
    "spearman": "Spearman Rank Correlation",
    "kendall": "Kendall’s τ Correlation",
    "logistic_regression": "Logistic Regression (binary outcome)",
    "multiple_regression": "Multiple Regression (with categorical encoding + VIF)",
    "posthoc": "Post-Hoc Pairwise Comparisons",
    "reliability": "Cronbach’s α (scale reliability)",
    "cramers_matrix": "Cramér’s V Association Matrix",
    "multinomial_logistic": "Multinomial Logistic Regression",
}

TITLE_KEYWORDS = [
    (re.compile(r"multinomial", re.I), "multinomial_logistic"),
    (re.compile(r"logistic regression", re.I), "logistic_regression"),
    (re.compile(r"multiple regression", re.I), "multiple_regression"),
    (re.compile(r"paired t-?test", re.I), "paired_ttest"),
    (re.compile(r"wilcoxon", re.I), "wilcoxon"),
    (re.compile(r"mcnemar", re.I), "mcnemar"),
    (re.compile(r"kruskal", re.I), "kruskal"),
    (re.compile(r"friedman", re.I), "friedman"),
    (re.compile(r"spearman", re.I), "spearman"),
    (re.compile(r"kendall", re.I), "kendall"),
    (re.compile(r"cram[ée]r", re.I), "cramers_matrix"),
    (re.compile(r"cross-?tab", re.I), "crosstab"),
    (re.compile(r"post-?hoc", re.I), "posthoc"),
    (re.compile(r"reliability|cronbach", re.I), "reliability"),
    (re.compile(r"anova", re.I), "anova"),
    (re.compile(r"t-?test", re.I), "ttest"),
    (re.compile(r"regression", re.I), "regression"),
    (re.compile(r"normality", re.I), "normality"),
    (re.compile(r"outlier", re.I), "outlier"),
    (re.compile(r"frequency", re.I), "frequency"),
    (re.compile(r"descriptive", re.I), "descriptive"),
    (re.compile(r"correlat", re.I), "correlation"),
]

MATRIX_TYPES = {"correlation", "spearman", "kendall", "cramers_matrix", "descriptive", "normality", "outlier", "reliability"}

# Two-variable stat types where a recovery with fewer than 2 columns is
# incomplete (a crosstab or t-test named after only one variable is missing
# its comparison variable, even if that one name matched cleanly).
TWO_VAR_TYPES = {"crosstab", "ttest", "anova", "kruskal", "mcnemar", "paired_ttest", "wilcoxon", "friedman"}


def min_columns_needed(stat_type):
    return 2 if stat_type in TWO_VAR_TYPES else 1


def normalize(s: str) -> str:
    return s.replace("–", "-").replace("—", "-").replace("‘", "'").replace("’", "'")


def guess_stat_type(title: str):
    if not title:
        return None
    nt = normalize(title)
    candidates = []
    for k, t in STAT_TITLES.items():
        nt2 = normalize(t)
        if nt == nt2 or nt.startswith(nt2 + " ") or nt.startswith(nt2 + "-") or nt.startswith(nt2 + " -"):
            candidates.append((k, nt2))
    if candidates:
        candidates.sort(key=lambda x: -len(x[1]))
        return candidates[0][0]
    for rx, k in TITLE_KEYWORDS:
        if rx.search(title):
            return k
    return None


def collect_known_columns(data: dict) -> set:
    cols = set()
    for t in data.get("tables", []) or []:
        cols.update(t.get("rows") or [])
        cols.update(t.get("columns") or [])
        for v in t.get("values") or []:
            if isinstance(v, dict) and v.get("field"):
                cols.add(v["field"])
        for f in (t.get("filters") or {}).keys():
            cols.add(f)
        cfg = t.get("_statConfig") or {}
        cols.update(cfg.get("columns") or [])
    for f in (data.get("columnTypeOverrides") or {}).keys():
        cols.add(f)
    for f in (data.get("projectFilters") or {}).keys():
        cols.add(f)
    return cols


def find_columns_in_title(title: str, known_cols: set):
    if not title:
        return []
    found = []
    for c in sorted(known_cols, key=len, reverse=True):
        if c and c in title and c not in found:
            found.append(c)
    return found


STOPWORDS = {
    "vs", "and", "the", "of", "by", "for", "in", "to", "on", "at", "with",
    "before", "after", "pre", "post", "test", "matrix", "correlation",
}


def title_remainder(title: str) -> str:
    """Strip the matched stat-type phrase off the front of a title, leaving
    just the analyst's own description of what it's about."""
    nt = normalize(title)
    for t in STAT_TITLES.values():
        nt2 = normalize(t)
        if nt.startswith(nt2):
            return nt[len(nt2):].strip(" -:–—")
    return title


def suggest_columns_by_keywords(title: str, known_cols: set, top_n: int = 3):
    """Fuzzy fallback when exact column names can't be found anywhere: score
    each known real column by how many meaningful words from the title's own
    description appear in it (whole-word match), and return the best guesses.
    These are SUGGESTIONS for a human to confirm, not an automatic answer —
    unlike resolve_table()'s columns, which are exact recoveries."""
    remainder = title_remainder(title)
    tokens = [w.lower() for w in re.findall(r"[A-Za-z]+", remainder) if len(w) >= 3 and w.lower() not in STOPWORDS]
    if not tokens:
        return []
    scored = []
    for c in known_cols:
        cl = c.lower()
        score = sum(1 for tok in tokens if re.search(r"\b" + re.escape(tok) + r"\b", cl))
        if score > 0:
            scored.append((score, c))
    scored.sort(key=lambda x: (-x[0], len(x[1])))
    return scored[:top_n]


def is_stat_table(t: dict) -> bool:
    return bool(t.get("_statResult")) or bool(t.get("_statResultData")) or (
        not t.get("rows") and not t.get("columns") and not t.get("values")
    )


def resolve_table(t: dict, known_cols: set) -> dict:
    title = t.get("title") or t.get("name") or ""
    cfg = t.get("_statConfig")
    if cfg and cfg.get("columns"):
        return {"status": "ok_via_statConfig", "statType": cfg.get("statType"), "columns": cfg["columns"]}
    stat_type = guess_stat_type(title)
    rd = t.get("_statResultData")
    if rd and stat_type in MATRIX_TYPES:
        matrix_cols = []
        for r in rd.get("rows") or []:
            if r and r[0] in known_cols and r[0] not in matrix_cols:
                matrix_cols.append(r[0])
        if matrix_cols:
            return {"status": "ok_via_resultData_matrix", "statType": stat_type, "columns": matrix_cols}
    title_cols = find_columns_in_title(title, known_cols)
    if title_cols:
        return {"status": "ok_via_title", "statType": stat_type, "columns": title_cols}
    return {"status": "unrecoverable_in_this_file", "statType": stat_type, "columns": []}


def load_headers_from_source(path: str) -> set:
    """Read just the header row of the real source data file — no row data —
    to widen the matching pool beyond what's still referenced by a table."""
    try:
        if path.lower().endswith((".xlsx", ".xls")):
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            wb.close()
            return {str(c).strip() for c in header_row if c is not None and str(c).strip()}
        else:
            import csv
            with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
                reader = csv.reader(f)
                header_row = next(reader)
            return {c.strip() for c in header_row if c and c.strip()}
    except Exception as e:  # noqa: BLE001
        print(f"could not read headers from {path}: {e}")
        return set()


def load_file(path: str):
    try:
        data = json.loads(Path(path).read_text())
    except Exception as e:  # noqa: BLE001
        return None, f"error reading {path}: {e}"
    if data.get("encrypted"):
        return None, f"{path}: password-protected — skipped"
    return data, None


def main():
    args = sys.argv[1:]
    source_xlsx = None
    if "--source-xlsx" in args:
        idx = args.index("--source-xlsx")
        source_xlsx = args[idx + 1]
        args = args[:idx] + args[idx + 2:]
    if len(args) < 1:
        print("usage: recover_stat_configs.py <current.tableforge> [other.tableforge ...] [--source-xlsx <path>]")
        return
    current_path, *other_paths = args
    data, err = load_file(current_path)
    if err:
        print(err)
        return

    known_cols = collect_known_columns(data)
    if source_xlsx:
        header_cols = load_headers_from_source(source_xlsx)
        print(f"source file headers: {len(header_cols)} columns read from {source_xlsx}")
        known_cols |= header_cols
    all_sources = []
    for op in other_paths:
        od, oerr = load_file(op)
        if oerr:
            print(oerr)
            continue
        known_cols |= collect_known_columns(od)
        for v in od.get("versions", []) or []:
            known_cols |= collect_known_columns(v)
            for t in v.get("tables", []) or []:
                all_sources.append((f"{op}@{v.get('saved_at')}", t))
        for t in od.get("tables", []) or []:
            all_sources.append((f"{op}@current", t))

    total_tables = len(data.get("tables") or [])
    stat_tables_seen = sum(1 for t in (data.get("tables") or []) if is_stat_table(t))
    print(f"{total_tables} total tables in current file, {stat_tables_seen} look like stat tables")
    print(f"known-columns pool: {len(known_cols)} real column names collected from every other table/version")
    print(f"{len(all_sources)} fallback source tables loaded from other files\n")

    results = []
    for t in data.get("tables", []) or []:
        if not is_stat_table(t):
            continue
        title = t.get("title") or t.get("name") or "(untitled)"
        r = resolve_table(t, known_cols)
        needed = min_columns_needed(r.get("statType"))
        if len(r["columns"]) < needed:
            # Current file's answer is missing or incomplete (e.g. a crosstab
            # recovered only 1 of its 2 variables from the title) — look for a
            # same-titled table elsewhere with a fuller answer before settling.
            best = r
            for label, ot in all_sources:
                ot_title = ot.get("title") or ot.get("name") or ""
                if ot_title != title:
                    continue
                orr = resolve_table(ot, known_cols)
                if len(orr["columns"]) > len(best["columns"]):
                    best = {"status": f"recovered_from:{label}", "statType": orr["statType"], "columns": orr["columns"]}
                    if len(best["columns"]) >= needed:
                        break
            if len(best["columns"]) > len(r["columns"]):
                r = best
            elif r["columns"]:
                r = {**r, "status": f"PARTIAL({len(r['columns'])}/{needed}):{r['status']}"}
        results.append((title, r))

    ok = [r for r in results if r[1]["columns"] and not str(r[1]["status"]).startswith("PARTIAL")]
    bad = [r for r in results if not r[1]["columns"] or str(r[1]["status"]).startswith("PARTIAL")]
    print(f"=== {len(results)} stat tables found | {len(ok)} recoverable | {len(bad)} still unrecoverable ===\n")
    for title, r in results:
        print(f"[{r['status']}] statType={r.get('statType')}\n    title: {title}")
        if r["columns"]:
            print(f"    columns: {', '.join(r['columns'])}\n")
            continue
        suggestions = suggest_columns_by_keywords(title, known_cols)
        if suggestions:
            print("    columns: NOT RECOVERED - best-guess candidates (confirm manually, not automatic):")
            for score, col in suggestions:
                print(f"      [{score} word match] {col}")
            print("")
        else:
            print("    columns: NOT RECOVERED - no keyword overlap found either; this one needs to be rebuilt from scratch\n")


if __name__ == "__main__":
    main()
