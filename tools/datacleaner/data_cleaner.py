"""
Data Cleaner — Lightweight web tool for interactive data cleaning.
Run:  python data_cleaner.py
Open:  http://localhost:5050
"""

import os, json, io, copy, math, pathlib
import threading, time, signal, sys
import requests as _requests
from flask import Flask, render_template, request, jsonify, send_file, g
import pandas as pd
import numpy as np
import pickle
from dotenv import load_dotenv

_root = pathlib.Path(__file__).resolve().parent.parent
load_dotenv(_root / ".env")
load_dotenv()  # also check current dir

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100 MB max upload

# Working copies are saved here — originals are NEVER modified
COPIES_DIR = pathlib.Path(__file__).resolve().parent / "working_copies"
COPIES_DIR.mkdir(exist_ok=True)

STATE_LOCK = threading.Lock()
STATE_DIRTY = False

# ── In-memory state ──────────────────────────────────────────────────────────
DATA = {
    "df": None,           # current DataFrame
    "original_df": None,  # snapshot for modification tracking on export
    "filename": None,     # original filename
    "copy_path": None,    # path to the working copy
    "sheet_name": None,   # active sheet name for Excel files
    "history": [],        # list of (description, DataFrame-copy) for undo
    "redo_stack": [],     # list of (description, DataFrame-copy) for redo
    "max_history": 20,
    "column_types": {},   # col_name -> type_name
    "active_filters": {},  # col_name -> {type, val1, val2}
}

def mark_state_dirty():
    global STATE_DIRTY
    STATE_DIRTY = True

def _save_state_loop():
    """Background thread that periodically saves state to disk if needed."""
    state_path = COPIES_DIR / "app_state.pkl"
    tmp_path = COPIES_DIR / "app_state_tmp.pkl"
    while True:
        time.sleep(2.0)
        global STATE_DIRTY
        if STATE_DIRTY:
            with STATE_LOCK:
                if not STATE_DIRTY:
                    continue
                try:
                    to_save = {
                        "df": DATA["df"],
                        "original_df": DATA["original_df"],
                        "filename": DATA["filename"],
                        "copy_path": DATA["copy_path"],
                        "sheet_name": DATA.get("sheet_name"),
                        "column_types": DATA["column_types"],
                        "active_filters": DATA["active_filters"]
                    }
                    with open(tmp_path, "wb") as f:
                        pickle.dump(to_save, f)
                    os.replace(tmp_path, state_path)
                    STATE_DIRTY = False
                except Exception as e:
                    print(f"Error saving state: {e}")

threading.Thread(target=_save_state_loop, daemon=True).start()

def signal_handler(sig, frame):
    """Ensure clean exit on Ctrl+C."""
    print("\n  Data Cleaner shutting down gracefully...")
    sys.exit(0)

signal.signal(signal.SIGINT, signal_handler)

def _save_state():
    """Legacy function call — just marks dirty now."""
    mark_state_dirty()

def _load_state():
    """Load state from disk on startup."""
    state_path = COPIES_DIR / "app_state.pkl"
    if state_path.exists():
        try:
            with open(state_path, "rb") as f:
                saved = pickle.load(f)
                DATA.update(saved)
                print(f"Restored state from {state_path}")
        except Exception as e:
            print(f"Error loading state: {e}")

def _push_undo(desc: str):
    """Snapshot current df for undo; clears redo stack."""
    if DATA["df"] is not None:
        DATA["history"].append((desc, DATA["df"].copy()))
        DATA["redo_stack"].clear()
        if len(DATA["history"]) > DATA["max_history"]:
            DATA["history"].pop(0)
    _save_state()

def _df():
    return DATA["df"]

def _get_filtered_mask(df, current_col):
    """Return a boolean mask for the current_col based on global filters."""
    mask = pd.Series(True, index=df.index)
    for col, f in DATA["active_filters"].items():
        if col not in df.columns: continue
        ftype = f.get("type")
        if ftype == "range":
            v_min, v_max = f.get("min"), f.get("max")
            # Convert to numeric for comparison if possible
            vals = pd.to_numeric(df[col], errors='coerce')
            if v_min is not None: mask &= (vals >= float(v_min))
            if v_max is not None: mask &= (vals <= float(v_max))
        elif ftype == "contain":
            val = str(f.get("value", ""))
            try:
                mask &= df[col].astype(str).str.contains(val, case=False, na=False, regex=False)
            except Exception:
                pass  # skip broken filter gracefully
        elif ftype == "missing":
            mask &= df[col].isna()
    
    # Virtual Filter: Row Completeness
    if "__missing_count" in DATA["active_filters"]:
        f = DATA["active_filters"]["__missing_count"]
        # Treat empty strings and 'nan' as missing
        is_missing = df.apply(lambda col: col.astype(str).str.strip().isin(['', 'nan', 'None', 'NaN']))
        missing_counts = is_missing.sum(axis=1)
        v_min, v_max = f.get("min"), f.get("max")
        if v_min is not None: mask &= (missing_counts >= float(v_min))
        if v_max is not None: mask &= (missing_counts <= float(v_max))
    
    # Global Search: Look in ALL columns
    if "__global_search" in DATA["active_filters"]:
        q = str(DATA["active_filters"]["__global_search"].get("value", "")).lower()
        if q:
            # Check if q is in any column for each row
            # We use astype(str).str.lower() for case-insensitive search
            global_mask = df.apply(lambda row: row.astype(str).str.lower().str.contains(q, regex=False).any(), axis=1)
            mask &= global_mask
        
    return mask

# ── Routes ───────────────────────────────────────────────────────────────────

@app.before_request
def acquire_lock():
    if request.path.startswith("/api/") and request.path != "/api/export":
        STATE_LOCK.acquire()
        g.lock_acquired = True

@app.teardown_request
def release_lock(exception=None):
    if getattr(g, 'lock_acquired', False):
        # Automatically mark state dirty at the end of every API request
        # if the request was successful and changed data
        mark_state_dirty()
        STATE_LOCK.release()

@app.route("/")
def index():
    return render_template("data_cleaner.html")

@app.route("/api/set_column_type", methods=["POST"])
def set_column_type():
    body = request.json
    col, dtype = body.get("column"), body.get("type")
    if col and dtype:
        DATA["column_types"][col] = dtype
        _save_state()
        return jsonify(ok=True)
    return jsonify(error="Missing column or type"), 400

@app.route("/api/set_filter", methods=["POST"])
def set_filter():
    body = request.json
    col, f = body.get("column"), body.get("filter")
    if col and f:
        DATA["active_filters"][col] = f
        _save_state()
        return jsonify(ok=True)
    return jsonify(error="Missing column or filter"), 400

@app.route("/api/clear_filter", methods=["POST"])
def clear_filter():
    col = request.json.get("column")
    if col in DATA["active_filters"]:
        del DATA["active_filters"][col]
        _save_state()
        return jsonify(ok=True)
    elif not col:
        DATA["active_filters"] = {}
        _save_state()
        return jsonify(ok=True)
    return jsonify(error="Filter not found"), 404


@app.route("/api/upload", methods=["POST"])
def upload():
    """Upload CSV or Excel file. A working copy is saved — original is never touched."""
    f = request.files.get("file")
    if not f:
        return jsonify(error="No file uploaded"), 400

    fname = f.filename.lower()
    raw_bytes = f.stream.read()

    # Save working copy
    copy_name = f"COPY_{f.filename}"
    copy_path = COPIES_DIR / copy_name
    with open(copy_path, "wb") as out:
        out.write(raw_bytes)

    # Parse from the saved copy
    try:
        if fname.endswith(".csv"):
            df = pd.read_csv(copy_path, dtype=str, keep_default_na=False, na_values=[''])
        elif fname.endswith((".xlsx", ".xls")):
            xls = pd.ExcelFile(copy_path, engine="openpyxl")
            sheet_names = xls.sheet_names
            if len(sheet_names) > 1:
                # Return and ask the user to select the correct sheet!
                return jsonify(ok=True, needs_sheet_selection=True, sheets=sheet_names, filename=f.filename, copy_path=copy_name)
            
            df = pd.read_excel(copy_path, sheet_name=sheet_names[0], engine="openpyxl", dtype=str, keep_default_na=False, na_values=[''])
            DATA["sheet_name"] = sheet_names[0]
        else:
            copy_path.unlink(missing_ok=True)
            return jsonify(error="Unsupported format. Use CSV or Excel."), 400
    except Exception as e:
        return jsonify(error=f"Parse error: {e}"), 400

    df.columns = df.columns.astype(str)
    DATA["df"] = df
    DATA["original_df"] = df.copy()
    DATA["filename"] = f.filename
    DATA["copy_path"] = str(copy_path)
    if fname.endswith(".csv"):
        DATA["sheet_name"] = None
    DATA["history"] = []
    _save_state()

    # For Excel files, detect if headers might be multi-row
    is_excel = fname.endswith((".xlsx", ".xls"))
    has_unnamed = sum(1 for c in df.columns if str(c).startswith("Unnamed")) > 3
    return jsonify(ok=True, rows=len(df), cols=len(df.columns),
                   columns=df.columns.tolist(), filename=f.filename,
                   copy_path=copy_name,
                   needs_header_config=is_excel and has_unnamed)

@app.route("/api/load_sheet", methods=["POST"])
def load_sheet():
    """Load a specific sheet after the user selected it from upload."""
    body = request.json
    copy_name = body.get("copy_path")
    sheet_name = body.get("sheet_name")
    filename = body.get("filename")
    
    if not copy_name or not sheet_name:
        return jsonify(error="Missing copy_path or sheet_name"), 400

    copy_path = COPIES_DIR / copy_name
    if not copy_path.exists():
        return jsonify(error="Working copy not found. Please upload again."), 400

    try:
        df = pd.read_excel(copy_path, sheet_name=sheet_name, engine="openpyxl", dtype=str, keep_default_na=False, na_values=[''])
    except Exception as e:
        return jsonify(error=f"Parse error: {e}"), 400

    df.columns = df.columns.astype(str)
    DATA["df"] = df
    DATA["original_df"] = df.copy()
    DATA["filename"] = filename or copy_name.replace("COPY_", "", 1)
    DATA["copy_path"] = str(copy_path)
    DATA["sheet_name"] = sheet_name
    DATA["history"] = []
    _save_state()

    has_unnamed = sum(1 for c in df.columns if str(c).startswith("Unnamed")) > 3
    return jsonify(ok=True, rows=len(df), cols=len(df.columns),
                   columns=df.columns.tolist(), filename=DATA["filename"],
                   copy_path=copy_name,
                   needs_header_config=has_unnamed)


# ── Server Files (shared uploads volume) ─────────────────────────────────────

UPLOADS_DIR = pathlib.Path(__file__).resolve().parent / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)
_MANIFEST = UPLOADS_DIR / "_manifest.json"

def _load_manifest():
    if _MANIFEST.exists():
        try:
            return json.loads(_MANIFEST.read_text())
        except Exception:
            pass
    return []

@app.route("/api/files")
def list_server_files():
    manifest = _load_manifest()
    valid = [e for e in manifest if (UPLOADS_DIR / e["stored_name"]).exists()]
    return jsonify(files=sorted(valid, key=lambda f: f.get("uploaded_at", ""), reverse=True))

@app.route("/api/files/<file_id>/load", methods=["POST"])
def load_server_file(file_id):
    manifest = _load_manifest()
    entry = next((f for f in manifest if f["id"] == file_id), None)
    if not entry:
        return jsonify(error="File not found on server"), 404
    src = UPLOADS_DIR / entry["stored_name"]
    if not src.exists():
        return jsonify(error="File missing from disk"), 404
    import shutil
    copy_name = f"COPY_{entry['original_name']}"
    copy_path = COPIES_DIR / copy_name
    shutil.copy2(str(src), str(copy_path))
    fname = entry["original_name"].lower()
    try:
        if fname.endswith(".csv") or fname.endswith(".tsv"):
            sep = "\t" if fname.endswith(".tsv") else ","
            df = pd.read_csv(copy_path, sep=sep, dtype=str, keep_default_na=False, na_values=[''])
        elif fname.endswith((".xlsx", ".xls")):
            xls = pd.ExcelFile(copy_path, engine="openpyxl")
            if len(xls.sheet_names) > 1:
                return jsonify(ok=True, needs_sheet_selection=True, sheets=xls.sheet_names,
                               filename=entry["original_name"], copy_path=copy_name)
            df = pd.read_excel(copy_path, sheet_name=xls.sheet_names[0], engine="openpyxl",
                               dtype=str, keep_default_na=False, na_values=[''])
            DATA["sheet_name"] = xls.sheet_names[0]
        else:
            return jsonify(error="Unsupported format"), 400
    except Exception as e:
        return jsonify(error=f"Parse error: {e}"), 400
    df.columns = df.columns.astype(str)
    DATA["df"] = df
    DATA["original_df"] = df.copy()
    DATA["filename"] = entry["original_name"]
    DATA["copy_path"] = str(copy_path)
    DATA["server_file_id"] = file_id
    DATA["history"] = []
    _save_state()
    return jsonify(ok=True, rows=len(df), cols=len(df.columns),
                   columns=df.columns.tolist(), filename=entry["original_name"],
                   copy_path=copy_name, server_file_id=file_id)


@app.route("/api/files/<file_id>/save-back", methods=["POST"])
def save_back_to_server(file_id):
    """Write the current cleaned DataFrame back to the shared uploads file."""
    df = DATA.get("df")
    if df is None:
        return jsonify(error="No data loaded"), 400
    manifest = _load_manifest()
    entry = next((f for f in manifest if f["id"] == file_id), None)
    if not entry:
        return jsonify(error="Server file not found"), 404
    dest = UPLOADS_DIR / entry["stored_name"]
    fname = entry["original_name"].lower()
    try:
        if fname.endswith((".xlsx", ".xls")):
            df.to_excel(str(dest), index=False, engine="openpyxl")
        elif fname.endswith(".tsv"):
            df.to_csv(str(dest), index=False, sep="\t")
        else:
            df.to_csv(str(dest), index=False)
    except Exception as e:
        return jsonify(error=f"Save failed: {e}"), 500
    entry["size_bytes"] = dest.stat().st_size
    entry["size_mb"] = round(entry["size_bytes"] / 1024 / 1024, 2)
    _MANIFEST.write_text(json.dumps(manifest, indent=2))
    return jsonify(ok=True, message=f"Saved cleaned data back to '{entry['original_name']}' on server")


@app.route("/api/preview_headers")
def preview_headers():
    """Show raw first N rows to help user identify header structure."""
    if not DATA["copy_path"]:
        return jsonify(error="No file loaded"), 400
    n = int(request.args.get("rows", 8))
    copy_path = pathlib.Path(DATA["copy_path"])
    fname = DATA["filename"].lower()
    if not fname.endswith((".xlsx", ".xls")):
        return jsonify(error="Header preview is only available for Excel files"), 400
    try:
        import openpyxl as xl
        wb = xl.load_workbook(copy_path, data_only=True)
        if DATA.get("sheet_name") and DATA["sheet_name"] in wb.sheetnames:
            ws = wb[DATA["sheet_name"]]
        else:
            ws = wb.active
        max_col = min(ws.max_column, 30)  # cap preview at 30 cols
        rows = []
        for r in range(1, n + 1):
            row = []
            for c in range(1, max_col + 1):
                v = ws.cell(row=r, column=c).value
                row.append(str(v) if v is not None else "")
            rows.append(row)
        # Also note merged ranges
        merges = [str(m) for m in ws.merged_cells.ranges]
        wb.close()
        return jsonify(ok=True, rows=rows, total_cols=ws.max_column,
                       total_rows=ws.max_row, merges_count=len(merges))
    except Exception as e:
        return jsonify(error=f"Preview error: {e}"), 400


@app.route("/api/configure_headers", methods=["POST"])
def configure_headers():
    """
    Re-parse the file using N header rows.
    Reads raw cells via openpyxl (handling merged cells) and builds
    composite column names like 'GENERAL INFORMATION > NAME OF BENEFICIARY'.
    """
    if not DATA["copy_path"]:
        return jsonify(error="No file loaded"), 400
    body = request.json
    header_rows = int(body.get("header_rows", 1))
    separator = body.get("separator", " > ")

    copy_path = pathlib.Path(DATA["copy_path"])
    fname = DATA["filename"].lower()

    try:
        if fname.endswith((".xlsx", ".xls")):
            import openpyxl as xl
            wb = xl.load_workbook(copy_path, data_only=True)
            if DATA.get("sheet_name") and DATA["sheet_name"] in wb.sheetnames:
                ws = wb[DATA["sheet_name"]]
            else:
                ws = wb.active

            # Build a mapping of merged cells → their top-left value
            merge_map = {}
            for mr in ws.merged_cells.ranges:
                top_val = ws.cell(row=mr.min_row, column=mr.min_col).value
                for r in range(mr.min_row, mr.max_row + 1):
                    for c in range(mr.min_col, mr.max_col + 1):
                        if (r, c) != (mr.min_row, mr.min_col):
                            merge_map[(r, c)] = top_val

            def get_cell(r, c):
                """Get cell value, resolving merged cells."""
                v = ws.cell(row=r, column=c).value
                if v is None and (r, c) in merge_map:
                    v = merge_map[(r, c)]
                return str(v).strip() if v is not None else ""

            # Build composite column names
            max_col = ws.max_column
            col_names = []
            for c in range(1, max_col + 1):
                parts = []
                for r in range(1, header_rows + 1):
                    val = get_cell(r, c)
                    if val and val not in parts:  # avoid duplicates
                        parts.append(val)
                name = separator.join(parts) if parts else f"Col_{c}"
                col_names.append(name)

            # Deduplicate column names
            seen = {}
            final_names = []
            for name in col_names:
                if name in seen:
                    seen[name] += 1
                    final_names.append(f"{name}_{seen[name]}")
                else:
                    seen[name] = 0
                    final_names.append(name)

            wb.close()

            # Read data using pandas with dtype=str to preserve raw appearance
            sheet_args = {"sheet_name": DATA["sheet_name"]} if DATA.get("sheet_name") else {}
            df = pd.read_excel(copy_path, skiprows=header_rows, header=None, dtype=str, engine="openpyxl", keep_default_na=False, na_values=[''], **sheet_args)
            
            # Align headers with actual data columns (handles pandas vs openpyxl mismatches)
            if len(final_names) > len(df.columns):
                final_names = final_names[:len(df.columns)]
            elif len(final_names) < len(df.columns):
                final_names += [f"Col_{i+1}" for i in range(len(final_names), len(df.columns))]
            
            df.columns = final_names
            
            # Drop fully empty rows
            df = df.dropna(how='all').reset_index(drop=True)

        elif fname.endswith(".csv"):
            df = pd.read_csv(copy_path, header=list(range(header_rows)), dtype=str, keep_default_na=False, na_values=[''])
            # Flatten MultiIndex columns
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = [separator.join(str(x) for x in col if str(x) != '' and not str(x).startswith('Unnamed'))
                              for col in df.columns.values]

        DATA["df"] = df
        DATA["original_df"] = df.copy()  # reset baseline after header config
        DATA["history"] = []  # clear history — new baseline after header config
        return jsonify(ok=True, rows=len(df), cols=len(df.columns),
                       columns=df.columns.tolist(),
                       header_rows=header_rows)
    except Exception as e:
        import traceback
        return jsonify(error=f"Header config error: {traceback.format_exc()}"), 400


@app.route("/api/info")
def info():
    """General dataset info."""
    df = _df()
    if df is None:
        return jsonify(error="No data loaded"), 400
    col_info = []
    for c in df.columns:
        dtype = str(df[c].dtype)
        total = len(df[c])
        missing = int(df[c].isna().sum())
        unique = int(df[c].nunique())
        stats = {}
        if pd.api.types.is_numeric_dtype(df[c]):
            s = df[c].dropna()
            if not s.empty:
                stats = {
                    "min": float(s.min()),
                    "max": float(s.max()),
                    "mean": round(float(s.mean()), 2),
                    "median": round(float(s.median()), 2),
                    "std": round(float(s.std()), 2) if len(s) > 1 else 0,
                    "q1": round(float(s.quantile(0.25)), 2),
                    "q3": round(float(s.quantile(0.75)), 2)
                }
            else:
                stats = {"min": None, "max": None, "mean": None}
        col_info.append(dict(name=c, dtype=dtype, total=total,
                             missing=missing, unique=unique, stats=stats))
    # Calculate row completeness distribution
    # Treat empty strings and 'nan' as missing
    is_missing = df.apply(lambda col: col.astype(str).str.strip().isin(['', 'nan', 'None', 'NaN']))
    missing_per_row = is_missing.sum(axis=1)
    
    buckets = {
        "0 (Complete)": int((missing_per_row == 0).sum()),
        "1 to 10 missing": int(((missing_per_row >= 1) & (missing_per_row <= 10)).sum()),
        "11 to 50 missing": int(((missing_per_row >= 11) & (missing_per_row <= 50)).sum()),
        "51+ missing": int((missing_per_row > 50).sum())
    }
    
    return jsonify(columns=col_info, rows=len(df), filename=DATA["filename"],
                   undo_count=len(DATA["history"]),
                   column_types=DATA["column_types"],
                   active_filters=DATA["active_filters"],
                   completeness_stats=buckets)


@app.route("/api/unique_values")
def unique_values():
    """Unique values + counts for a column."""
    df = _df()
    if df is None:
        return jsonify(error="No data loaded"), 400
    col = request.args.get("column")
    if col not in df.columns:
        return jsonify(error=f"Column '{col}' not found"), 400

    # Respect global filters (except maybe the one on this column?)
    # Usually users want to see the effect of OTHER filters on this column.
    # But if they filtered THIS column to "2026-01-23", they only want to see that.
    mask = _get_filtered_mask(df, None)
    filtered_df = df[mask]
    
    vc = filtered_df[col].value_counts(dropna=False)
    items = []
    for val, cnt in vc.items():
        display = "(missing)" if pd.isna(val) else str(val)
        items.append({"value": display, "raw": None if pd.isna(val) else val,
                       "count": int(cnt),
                       "pct": round(cnt / len(filtered_df) * 100, 1) if len(filtered_df) > 0 else 0})
    missing = int(filtered_df[col].isna().sum())
    
    # Also return if this column has an active filter or declared type
    active_filter = DATA["active_filters"].get(col)
    column_type = DATA["column_types"].get(col, "Text")
    
    return jsonify(column=col, total=len(df), missing=missing,
                   unique=int(df[col].nunique()), items=items,
                   active_filter=active_filter, column_type=column_type)


@app.route("/api/replace", methods=["POST"])
def replace_values():
    """Replace values in a column."""
    df = _df()
    if df is None:
        return jsonify(error="No data loaded"), 400
    body = request.json
    col = body.get("column")
    find_val = body.get("find")
    replace_val = body.get("replace")
    use_regex = body.get("regex", False)

    if col not in df.columns:
        return jsonify(error=f"Column '{col}' not found"), 400

    _push_undo(f"Replace '{find_val}' → '{replace_val}' in {col}")
    
    # Respect filters
    mask = _get_filtered_mask(df, col)

    if find_val == "(missing)":
        mask &= df[col].isna()
        count = int(mask.sum())
        df.loc[mask, col] = replace_val
    else:
        if use_regex:
            match_mask = df[col].astype(str).str.contains(find_val, regex=True, na=False)
            final_mask = mask & match_mask
            count = int(final_mask.sum())
            # For regex replace, we need to be careful with Series vs DataFrame
            df.loc[final_mask, col] = df.loc[final_mask, col].astype(str).str.replace(find_val, replace_val, regex=True)
        else:
            match_mask = df[col].astype(str) == str(find_val)
            final_mask = mask & match_mask
            count = int(final_mask.sum())
            df.loc[final_mask, col] = replace_val

    DATA["df"] = df
    return jsonify(ok=True, replaced=count)


@app.route("/api/trim_whitespace", methods=["POST"])
def trim_whitespace():
    """Trim whitespace from selected columns."""
    df = _df()
    if df is None:
        return jsonify(error="No data loaded"), 400
    body = request.json
    columns = body.get("columns", [])
    if not columns:
        columns = [c for c in df.columns if df[c].dtype == object]

    _push_undo(f"Trim whitespace: {columns}")
    
    # Respect filters
    mask = _get_filtered_mask(df, None)
    
    trimmed = 0
    for c in columns:
        if c in df.columns and df[c].dtype == object:
            subset = df.loc[mask, c]
            # Only strip string values, skip NaN
            stripped = subset.where(subset.isna(), subset.astype(str).str.strip())
            changed = (subset.fillna("__NA__") != stripped.fillna("__NA__"))
            trimmed += int(changed.sum())
            df.loc[mask, c] = stripped
    DATA["df"] = df
    return jsonify(ok=True, trimmed=trimmed)


@app.route("/api/round_values", methods=["POST"])
def round_values():
    """Round numeric values in a column."""
    df = _df()
    if df is None:
        return jsonify(error="No data loaded"), 400
    body = request.json
    col = body.get("column")
    decimals = int(body.get("decimals", 0))

    if col not in df.columns:
        return jsonify(error=f"Column '{col}' not found"), 400

    _push_undo(f"Round {col} to {decimals} decimals")
    
    # Respect filters
    mask = _get_filtered_mask(df, col)
    
    # Convert to numeric, round, then back to string (preserving NaN)
    vals = pd.to_numeric(df.loc[mask, col], errors='coerce')
    rounded = vals.round(decimals)
    
    # Only update where rounding was possible
    df.loc[mask, col] = rounded.astype(str).replace('nan', np.nan)
    
    DATA["df"] = df
    return jsonify(ok=True, modified=int(mask.sum()))


@app.route("/api/cross_validate", methods=["POST"])
def cross_validate():
    """Check a rule between two columns (e.g., Col1 > Col2)."""
    df = _df()
    if df is None:
        return jsonify(error="No data loaded"), 400
    body = request.json
    col1 = body.get("col1")
    col2 = body.get("col2")
    op = body.get("operator")  # '>', '<', '==', '!=', '>=', '<='

    if col1 not in df.columns or col2 not in df.columns:
        return jsonify(error="One or both columns not found"), 400

    # Respect filters
    mask = _get_filtered_mask(df, None)
    subset = df[mask]

    # Try to convert to numeric for comparison if numeric operators are used
    if op in ['>', '<', '>=', '<=']:
        v1 = pd.to_numeric(subset[col1], errors='coerce')
        v2 = pd.to_numeric(subset[col2], errors='coerce')
    else:
        v1 = subset[col1].astype(str)
        v2 = subset[col2].astype(str)

    # Perform comparison
    if op == '>': violations_mask = v1 <= v2
    elif op == '<': violations_mask = v1 >= v2
    elif op == '>=': violations_mask = v1 < v2
    elif op == '<=': violations_mask = v1 > v2
    elif op == '==': violations_mask = v1 != v2
    elif op == '!=': violations_mask = v1 == v2
    else: return jsonify(error="Invalid operator"), 400

    # Drop NaNs from violations (if we can't compare, it's not a violation?)
    # Or should we flag comparisons with NaN as violations? 
    # Usually, we only care about valid values that break the rule.
    violations_mask &= v1.notna() & v2.notna()
    
    violators = subset[violations_mask]
    results = []
    for idx, row in violators.head(500).iterrows(): # cap at 500
        results.append({
            "row": idx,
            "v1": row[col1],
            "v2": row[col2]
        })

    return jsonify(ok=True, 
                   total_checked=int(mask.sum()),
                   total_violations=int(violations_mask.sum()),
                   violations=results)


@app.route("/api/fill_proportional", methods=["POST"])
def fill_proportional():
    """
    Fill missing values using proportional distribution.
    Distributes missing cells across existing value frequencies,
    maintaining the current distribution pattern.
    """
    df = _df()
    if df is None:
        return jsonify(error="No data loaded"), 400
    body = request.json
    col = body.get("column")
    if col not in df.columns:
        return jsonify(error=f"Column '{col}' not found"), 400

    missing_count = int(df[col].isna().sum())
    if missing_count == 0:
        return jsonify(ok=True, filled=0, message="No missing values")

    non_null = df[col].dropna()
    if len(non_null) == 0:
        return jsonify(error="Column is entirely empty, cannot distribute"), 400

    _push_undo(f"Proportional fill: {col}")

    vc = non_null.value_counts(normalize=True)

    # Calculate how many of each value to assign
    allocations = (vc * missing_count).apply(math.floor).astype(int)
    remainder = missing_count - allocations.sum()

    # Distribute remainder to top values
    fractional = (vc * missing_count) - allocations
    top_remainders = fractional.nlargest(remainder).index
    for v in top_remainders:
        allocations[v] += 1

    # Build fill list
    fill_values = []
    for val, cnt in allocations.items():
        fill_values.extend([val] * cnt)

    np.random.shuffle(fill_values)

    missing_mask = df[col].isna()
    # Respect filters
    active_mask = _get_filtered_mask(df, col)
    final_missing_mask = missing_mask & active_mask

    missing_idx = df[final_missing_mask].index.tolist()
    if not missing_idx:
        return jsonify(ok=True, filled=0, message="No missing values in current filter")

    actual_filled = 0
    for i, idx in enumerate(missing_idx):
        if i < len(fill_values):
            df.at[idx, col] = fill_values[i]
            actual_filled += 1

    DATA["df"] = df

    # Show resulting distribution
    new_vc = df[col].value_counts(dropna=False)
    distribution = {str(k): int(v) for k, v in new_vc.items()}
    return jsonify(ok=True, filled=actual_filled, distribution=distribution)


@app.route("/api/ai_correct", methods=["POST"])
def ai_correct():
    """Send column data to Gemini for AI-powered correction suggestions."""
    df = _df()
    if df is None:
        return jsonify(error="No data loaded"), 400
    body = request.json
    columns = body.get("columns", [])
    context_columns = body.get("context", [])
    custom_prompt = body.get("prompt", "")

    if not columns:
        return jsonify(error="No columns selected"), 400

    all_req_cols = list(set(columns + context_columns))
    for c in all_req_cols:
        if c not in df.columns:
            return jsonify(error=f"Column '{c}' not found"), 400

    # Respect filters for sampling
    active_mask = _get_filtered_mask(df, None)
    filtered_df = df[active_mask]
    
    if len(filtered_df) == 0:
        return jsonify(error="No rows match the current filter"), 400

    data_to_send = {}
    is_row_level = len(context_columns) > 0

    if is_row_level:
        # Send a sample of rows containing both target and context
        sample_size = 50
        missing_any = filtered_df[columns].isna().any(axis=1)
        prioritized = filtered_df[missing_any].head(sample_size)
        if len(prioritized) < sample_size:
            extra = filtered_df[~missing_any].head(sample_size - len(prioritized))
            prioritized = pd.concat([prioritized, extra])
        
        sample_data = prioritized[all_req_cols].to_dict(orient="records")
        data_to_send = {"type": "row_sample", "rows": sample_data}
    else:
        # Unique value counts
        col_data = {}
        for c in columns:
            vc = filtered_df[c].value_counts(dropna=False).head(100)
            items = {}
            for val, cnt in vc.items():
                key = "(missing)" if pd.isna(val) else str(val)
                items[key] = int(cnt)
            col_data[c] = items
        data_to_send = {"type": "unique_counts", "data": col_data}

    prompt = f"""You are a data cleaning expert. Analyze the following data from a dataset and suggest corrections.

CONTENT TYPE: {data_to_send['type']}
DATA:
{json.dumps(data_to_send, indent=2, ensure_ascii=False)}

{f"ADDITIONAL CONTEXT/INSTRUCTIONS: {custom_prompt}" if custom_prompt else ""}

INSTRUCTIONS:
1. Identify typos, inconsistencies, formatting issues, and incorrect values.
2. If context columns are provided, use them to infer or validate values in the target columns.
3. Suggest specific replacements (original value → corrected value).
4. Note that '(missing)' or null values should be filled if the context allows.

Respond in this EXACT JSON format:
{{
  "suggestions": [
    {{
      "column": "column_name",
      "find": "original_value",
      "replace": "corrected_value",
      "reason": "brief reason"
    }}
  ],
  "notes": "any general observations about the data quality"
}}

Return ONLY valid JSON, no markdown fences."""

    try:
        import google.generativeai as genai
        api_key = os.getenv("GOOGLE_API_KEY")
        if not api_key:
            return jsonify(error="GOOGLE_API_KEY not set in .env"), 400

        genai.configure(api_key=api_key)
        model = genai.GenerativeModel("gemini-3-flash-preview")
        response = model.generate_content(prompt)
        text = response.text.strip()
        # Strip markdown code fences (```json ... ``` or ``` ... ```)
        if text.startswith("```"):
            # Remove opening fence line
            text = text.split("\n", 1)[1] if "\n" in text else text[3:]
        if text.endswith("```"):
            text = text[:-3]
        text = text.strip()

        result = json.loads(text)
        return jsonify(ok=True, result=result)
    except json.JSONDecodeError:
        return jsonify(ok=True, result={"suggestions": [], "notes": text})
    except Exception as e:
        return jsonify(error=f"AI error: {str(e)}"), 500


@app.route("/api/apply_ai_suggestions", methods=["POST"])
def apply_ai_suggestions():
    """Apply selected AI suggestions."""
    df = _df()
    if df is None:
        return jsonify(error="No data loaded"), 400
    body = request.json
    suggestions = body.get("suggestions", [])
    if not suggestions:
        return jsonify(error="No suggestions to apply"), 400

    _push_undo("Apply AI suggestions")
    
    # Respect filters
    active_mask = _get_filtered_mask(df, None) # global filter
    
    applied = 0
    for s in suggestions:
        col = s.get("column")
        find = s.get("find")
        replace = s.get("replace")
        if col in df.columns:
            if find == "(missing)":
                target_mask = df[col].isna() & active_mask
                count = int(target_mask.sum())
                df.loc[target_mask, col] = replace
            else:
                target_mask = (df[col].astype(str) == str(find)) & active_mask
                count = int(target_mask.sum())
                df.loc[target_mask, col] = replace
            applied += count

    DATA["df"] = df
    return jsonify(ok=True, applied=applied)
 
 
@app.route("/api/manual_edit", methods=["POST"])
def manual_edit():
    """Manually update a single cell."""
    df = _df()
    if df is None:
        return jsonify(error="No data loaded"), 400
    body = request.json
    row_idx = body.get("row_idx")
    col = body.get("column")
    new_val = body.get("value")
 
    if col not in df.columns:
        return jsonify(error=f"Column '{col}' not found"), 400
    
    try:
        # We use loc with the dataframe index to be precise
        # (row_idx should be the index label from display_df)
        _push_undo(f"Manual edit in {col}")
        df.at[row_idx, col] = new_val
        DATA["df"] = df
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(error=f"Update failed: {str(e)}"), 400


@app.route("/api/generate_regex", methods=["POST"])
def generate_regex():
    """Convert a plain English description into a regular expression using AI."""
    body = request.json
    description = body.get("description", "")
    if not description:
        return jsonify(error="No description provided"), 400

    prompt = (
        r"You are a regex expert. Convert the following description of a text pattern "
        r"into a Python-compatible Regular Expression (regex).\n\n"
        f"DESCRIPTION: {description}\n\n"
        "INSTRUCTIONS:\n"
        "1. Provide ONLY the raw regex string.\n"
        "2. Do not include quotes, markdown fences, or any explanation.\n"
        "3. Ensure it works with the Python `re` module.\n"
        "4. If the description is vague, provide the most likely pattern.\n\n"
        'Example: "10 digit phone number" -> ^\\d{10}$\n'
        'Example: "starts with GK" -> ^GK.*\n\n'
        "REGEX:"
    )

    try:
        import google.generativeai as genai
        api_key = os.getenv("GOOGLE_API_KEY")
        if not api_key:
            return jsonify(error="GOOGLE_API_KEY not set in .env"), 400

        genai.configure(api_key=api_key)
        model = genai.GenerativeModel("gemini-3-flash-preview")
        response = model.generate_content(prompt)
        regex = response.text.strip().replace("`", "")
        return jsonify(ok=True, regex=regex)
    except Exception as e:
        return jsonify(error=f"AI error: {str(e)}"), 500


@app.route("/api/preview")
def preview():
    """Paginated data preview."""
    df = _df()
    if df is None:
        return jsonify(error="No data loaded"), 400

    # Respect filters in preview
    active_mask = _get_filtered_mask(df, None)
    display_df = df[active_mask]
    
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 50))
    start = (page - 1) * per_page
    end = start + per_page
    subset = display_df.iloc[start:end]

    rows = []
    for idx, row in subset.iterrows():
        r = {"__row_idx": idx}
        for c in df.columns:
            v = row[c]
            r[c] = None if pd.isna(v) else v
            # Convert numpy types to native Python types for JSON
            if isinstance(r[c], (np.integer,)):
                r[c] = int(r[c])
            elif isinstance(r[c], (np.floating,)):
                r[c] = float(r[c])
        rows.append(r)

    return jsonify(columns=df.columns.tolist(), rows=rows,
                   page=page, per_page=per_page, total=len(display_df),
                   total_pages=math.ceil(len(display_df) / per_page))

# ── Feature A: Categorical Mapping ───────────────────────────────────────────

@app.route("/api/categorical_map", methods=["POST"])
def categorical_map():
    """Bulk replace values using a mapping table."""
    df = _df()
    if df is None:
        return jsonify(error="No data loaded"), 400
    body = request.json
    col = body.get("column")
    mappings = body.get("mappings", {})
    if col not in df.columns:
        return jsonify(error=f"Column '{col}' not found"), 400
    if not mappings:
        return jsonify(error="No mappings provided"), 400

    _push_undo(f"Categorical map on '{col}'")
    
    # Respect filters
    active_mask = _get_filtered_mask(df, col)
    
    count = 0
    for old_val, new_val in mappings.items():
        if old_val == "(missing)":
            mask = df[col].isna() & active_mask
            count += int(mask.sum())
            df.loc[mask, col] = new_val
        else:
            mask = (df[col].astype(str) == old_val) & active_mask
            count += int(mask.sum())
            df.loc[mask, col] = new_val
    DATA["df"] = df
    return jsonify(ok=True, replaced=count)


# ── Feature B: Date Standardizer ─────────────────────────────────────────────

@app.route("/api/standardize_dates", methods=["POST"])
def standardize_dates():
    """Auto-detect date formats and convert to a standard format."""
    df = _df()
    if df is None:
        return jsonify(error="No data loaded"), 400
    body = request.json
    col = body.get("column")
    target_fmt = body.get("format", "%d-%m-%Y")  # DD-MM-YYYY default
    dayfirst = body.get("dayfirst", True)

    if col not in df.columns:
        return jsonify(error=f"Column '{col}' not found"), 400

    _push_undo(f"Standardize dates in '{col}'")

    # Respect filters
    active_mask = _get_filtered_mask(df, col)

    original = df[col].copy()
    parsed = pd.to_datetime(df[col], dayfirst=dayfirst, errors='coerce')

    # Vectorized: only rows that are filtered-in, non-null original, and parseable
    has_value = original.notna() & active_mask
    parseable = has_value & parsed.notna()
    unparseable = has_value & parsed.isna()

    # Format all parseable dates at once
    formatted = parsed[parseable].dt.strftime(target_fmt)
    df.loc[parseable, col] = formatted

    converted = int(parseable.sum())
    failed = int(unparseable.sum())

    DATA["df"] = df
    return jsonify(ok=True, converted=converted, failed=failed,
                   total=int(has_value.sum()))


# ── Feature C: Outlier Detection ─────────────────────────────────────────────

@app.route("/api/detect_outliers", methods=["POST"])
def detect_outliers():
    """Find outliers using Z-score for numeric columns."""
    df = _df()
    if df is None:
        return jsonify(error="No data loaded"), 400
    body = request.json
    col = body.get("column")
    threshold = float(body.get("threshold", 3.0))

    if col not in df.columns:
        return jsonify(error=f"Column '{col}' not found"), 400

    # Convert to numeric
    numeric = pd.to_numeric(df[col], errors='coerce')
    valid = numeric.dropna()
    if len(valid) < 3:
        return jsonify(error="Not enough numeric values for outlier detection"), 400

    mean = float(valid.mean())
    std = float(valid.std())
    if std == 0:
        return jsonify(ok=True, outliers=[], mean=mean, std=0,
                       message="All values are the same — no outliers possible")

    outliers = []
    for idx in valid.index:
        z = (valid[idx] - mean) / std
        if abs(z) > threshold:
            outliers.append({
                "row": int(idx) + 1,  # 1-indexed for display
                "value": float(valid[idx]),
                "z_score": round(z, 2)
            })

    return jsonify(ok=True, outliers=outliers, mean=round(mean, 2),
                   std=round(std, 2), threshold=threshold,
                   total_checked=len(valid))


@app.route("/api/validate_format", methods=["POST"])
def validate_format():
    """Validate column values against a regex pattern."""
    import re as re_mod
    df = _df()
    if df is None:
        return jsonify(error="No data loaded"), 400
    body = request.json
    col = body.get("column")
    pattern = body.get("pattern", "")

    if col not in df.columns:
        return jsonify(error=f"Column '{col}' not found"), 400
    if not pattern:
        return jsonify(error="No pattern provided"), 400

    try:
        regex = re_mod.compile(pattern)
    except re_mod.error as e:
        return jsonify(error=f"Invalid regex: {e}"), 400

    non_null = df[col].dropna()
    mismatches = {}
    for idx, val in non_null.items():
        s = str(val)
        if not regex.fullmatch(s):
            if s in mismatches:
                mismatches[s]["count"] += 1
                if len(mismatches[s]["rows"]) < 5:
                    mismatches[s]["rows"].append(int(idx) + 1)
            else:
                mismatches[s] = {"count": 1, "rows": [int(idx) + 1]}

    items = [{"value": k, "count": v["count"], "sample_rows": v["rows"]}
             for k, v in sorted(mismatches.items(), key=lambda x: -x[1]["count"])]

    return jsonify(ok=True, mismatches=items[:100],
                   total_checked=len(non_null),
                   total_mismatched=sum(v["count"] for v in mismatches.values()),
                   pattern=pattern)


@app.route("/api/undo", methods=["POST"])
def undo():
    """Undo last operation, pushing current state to redo stack."""
    if not DATA["history"]:
        return jsonify(error="Nothing to undo"), 400
    # Save current state to redo stack before undoing
    if DATA["df"] is not None:
        DATA["redo_stack"].append(("(before undo)", DATA["df"].copy()))
    desc, prev_df = DATA["history"].pop()
    DATA["df"] = prev_df
    _save_state()
    return jsonify(ok=True, undone=desc, remaining=len(DATA["history"]), redo_count=len(DATA["redo_stack"]))


@app.route("/api/redo", methods=["POST"])
def redo():
    """Redo last undone operation."""
    if not DATA["redo_stack"]:
        return jsonify(error="Nothing to redo"), 400
    # Save current state to undo history
    if DATA["df"] is not None:
        DATA["history"].append(("(before redo)", DATA["df"].copy()))
    desc, next_df = DATA["redo_stack"].pop()
    DATA["df"] = next_df
    _save_state()
    return jsonify(ok=True, redone=desc, remaining=len(DATA["redo_stack"]), undo_count=len(DATA["history"]))


@app.route("/api/export")
def export():
    """Export cleaned data as CSV or Excel.
    Excel exports include modification tracking: yellow background + comments.
    """
    df = _df()
    if df is None:
        return jsonify(error="No data loaded"), 400
    fmt = request.args.get("format", "csv")
    orig_name = DATA["filename"] or "cleaned_data"
    base = os.path.splitext(orig_name)[0]

    if fmt == "excel":
        import openpyxl as xl
        from openpyxl.styles import PatternFill
        from openpyxl.comments import Comment

        out_name = f"{base}_cleaned.xlsx"
        buf = io.BytesIO()

        orig_df = DATA.get("original_df")
        include_comments = request.args.get("comments", "true").lower() == "true"

        if orig_df is not None and include_comments:
            # Tracked export with yellow highlights and "Was: ..." comments
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                df.to_excel(writer, index=False)
                wb = writer.book
                ws = writer.sheets["Sheet1"]
                yellow = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

                common_cols = [c for c in df.columns if c in orig_df.columns]
                if common_cols:
                    min_rows = min(len(df), len(orig_df))
                    cur_subset = df[common_cols].iloc[:min_rows]
                    orig_subset = orig_df[common_cols].iloc[:min_rows]
                    
                    cur_mask = cur_subset.isna()
                    orig_mask = orig_subset.isna()
                    diff_mask = (cur_mask != orig_mask) | (~cur_mask & ~orig_mask & (cur_subset.astype(str) != orig_subset.astype(str)))
                    
                    diff_series = diff_mask.stack()
                    diff_coords = diff_series[diff_series].index
                    col_to_idx = {c: i + 1 for i, c in enumerate(df.columns)}
                    
                    for ri, cname in diff_coords:
                        ci_excel = col_to_idx[cname]
                        ri_excel = ri + 2  # header + 1-index
                        orig_val = orig_subset.at[ri, cname]
                        cell = ws.cell(row=ri_excel, column=ci_excel)
                        cell.fill = yellow
                        was = "(empty)" if pd.isna(orig_val) else str(orig_val)
                        cell.comment = Comment(f"Was: {was}", "Data Cleaner")
        else:
            # Plain export (no tracking)
            df.to_excel(buf, index=False, engine="openpyxl")

        buf.seek(0)
        # Also save to working_copies on disk
        disk_path = COPIES_DIR / out_name
        try:
            with open(disk_path, "wb") as f:
                f.write(buf.getvalue())
        except Exception:
            pass # ignore disk save errors for now

        return send_file(buf, as_attachment=True,
                         download_name=out_name,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        out_name = f"{base}_cleaned.csv"
        buf = io.BytesIO()
        df.to_csv(buf, index=False)
        buf.seek(0)
        # Also save to working_copies
        disk_path = COPIES_DIR / out_name
        df.to_csv(disk_path, index=False)
        return send_file(buf, as_attachment=True,
                         download_name=out_name,
                         mimetype="text/csv")


@app.route("/api/fill_na", methods=["POST"])
def fill_na():
    df = _df()
    if df is None: return jsonify(error="No data loaded"), 400
    col = request.json.get("column")
    if col not in df.columns: return jsonify(error="Column not found"), 400
    
    _push_undo(f"Fill missing with 'NA' in {col}")
    mask = df[col].isna() & _get_filtered_mask(df, col)
    count = int(mask.sum())
    if count > 0:
        df.loc[mask, col] = "NA"
        DATA["df"] = df
    return jsonify(ok=True, filled=count)



@app.route("/api/remove_spaces", methods=["POST"])
def remove_spaces():
    df = _df()
    if df is None: return jsonify(error="No data loaded"), 400
    col = request.json.get("column")
    if col not in df.columns: return jsonify(error="Column not found"), 400
    
    _push_undo(f"Remove all spaces in {col}")
    mask = _get_filtered_mask(df, col)
    
    subset = df.loc[mask, col]
    # Remove spaces safely, only from non-null string values
    no_spaces = subset.where(subset.isna(), subset.astype(str).str.replace(r"\s+", "", regex=True))
    changed = (subset.fillna("__NA__") != no_spaces.fillna("__NA__"))
    count = int(changed.sum())
    
    if count > 0:
        df.loc[mask, col] = no_spaces
        DATA["df"] = df
    return jsonify(ok=True, modified=count)


@app.route("/api/text_case", methods=["POST"])
def text_case():
    df = _df()
    if df is None: return jsonify(error="No data loaded"), 400
    body = request.json
    col = body.get("column")
    case_type = body.get("case_type")
    
    if col not in df.columns: return jsonify(error="Column not found"), 400
    if case_type not in ["upper", "lower", "proper"]: return jsonify(error="Invalid case type"), 400
    
    _push_undo(f"Convert {col} to {case_type} case")
    mask = _get_filtered_mask(df, col)
    
    subset = df.loc[mask, col]
    
    if case_type == "upper":
        transformed = subset.where(subset.isna(), subset.astype(str).str.upper())
    elif case_type == "lower":
        transformed = subset.where(subset.isna(), subset.astype(str).str.lower())
    elif case_type == "proper":
        transformed = subset.where(subset.isna(), subset.astype(str).str.title())
        
    changed = (subset.fillna("__NA__") != transformed.fillna("__NA__"))
    count = int(changed.sum())
    
    if count > 0:
        df.loc[mask, col] = transformed
        DATA["df"] = df
    return jsonify(ok=True, modified=count)



@app.route("/api/scan_errors", methods=["POST"])
def scan_errors():
    """Find values that don't match the declared column type."""
    df = _df()
    if df is None:
        return jsonify(error="No data loaded"), 400
    body = request.json
    col = body.get("column")
    dtype = body.get("type") or DATA["column_types"].get(col, "Text")
    
    if col not in df.columns:
        return jsonify(error=f"Column '{col}' not found"), 400
    
    errors = []
    series = df[col].astype(str).replace("nan", "")
    
    if dtype == "Number":
        # values that are not empty and not numeric
        mask = (series != "") & pd.to_numeric(series, errors='coerce').isna()
    elif dtype == "Date":
        # values that are not empty and not parseable as date
        # we use a very lenient dayfirst=True for India
        mask = (series != "") & pd.to_datetime(series, errors='coerce', dayfirst=True).isna()
    elif dtype == "Phone":
        # 10 digits
        mask = (series != "") & (~series.str.match(r"^\d{10}$", na=False))
    elif dtype == "Aadhaar":
        # 12 digits
        mask = (series != "") & (~series.str.match(r"^\d{12}$", na=False))
    else:
        # Text always matches
        mask = pd.Series(False, index=df.index)

    err_df = df[mask].copy()
    vc = err_df[col].value_counts()
    for val, cnt in vc.items():
        errors.append({"value": str(val), "count": int(cnt)})
        
    return jsonify(ok=True, errors=errors, total_errors=int(mask.sum()))


@app.route("/api/apply_recovery", methods=["POST"])
def apply_recovery():
    """Apply the recovered mapping from recovered_mapping.json."""
    df = _df()
    if df is None:
        return jsonify(error="No data loaded"), 400
    
    body = request.json
    col = body.get("column")
    if col not in df.columns:
        return jsonify(error=f"Column '{col}' not found"), 400
    
    recovery_path = pathlib.Path(__file__).resolve().parent / "recovered_mapping.json"
    if not recovery_path.exists():
        return jsonify(error="Recovery file not found"), 404
    
    try:
        with open(recovery_path, "r") as f:
            mapping = json.load(f)
        
        _push_undo("Apply recovery mapping")
        
        # Respect filters
        active_mask = _get_filtered_mask(df, col)
        
        # Apply the mapping
        count = 0
        for old_val, new_val in mapping.items():
            mask = (df[col] == old_val) & active_mask
            if mask.any():
                df.loc[mask, col] = new_val
                count += int(mask.sum())
        
        _save_state()
        return jsonify(ok=True, message=f"Successfully reapplied {count} mappings to '{col}'")
    except Exception as e:
        return jsonify(error=f"Recovery error: {e}"), 500


# ── FieldGovern Integration + proxy endpoints ────────────────────────────────

@app.route("/api/fg/programs", methods=["POST"])
def proxy_fg_programs():
    body = request.json or {}
    fg_base_url = body.get("fg_base_url", "").rstrip("/")
    token = body.get("token", "")
    if not fg_base_url or not token:
        return jsonify(error="fg_base_url and token required"), 400
    try:
        resp = _requests.get(f"{fg_base_url}/api/v1/programs/",
                             headers={"Authorization": f"Bearer {token}"}, timeout=30)
        return jsonify(resp.json()), resp.status_code
    except Exception as e:
        return jsonify(error=str(e)), 502


@app.route("/api/fg/questionnaires", methods=["POST"])
def proxy_fg_questionnaires():
    body = request.json or {}
    fg_base_url = body.get("fg_base_url", "").rstrip("/")
    token = body.get("token", "")
    program_id = body.get("program_id", "")
    if not fg_base_url or not token or not program_id:
        return jsonify(error="fg_base_url, token and program_id required"), 400
    try:
        resp = _requests.get(f"{fg_base_url}/api/v1/programs/{program_id}/questionnaires",
                             headers={"Authorization": f"Bearer {token}"}, timeout=30)
        return jsonify(resp.json()), resp.status_code
    except Exception as e:
        return jsonify(error=str(e)), 502


@app.route("/api/save-to-account", methods=["POST"])
def save_to_account():
    """Save current cleaned data to the user's FG account as a tool project."""
    body = request.json or {}
    fg_base_url = body.get("fg_base_url", "").rstrip("/")
    token = body.get("token", "")
    name = body.get("name", "").strip() or "Cleaner Export"
    program_id = body.get("program_id") or None
    if not fg_base_url or not token:
        return jsonify(error="fg_base_url and token required"), 400
    df = _df()
    if df is None:
        return jsonify(error="No data loaded"), 400
    # Export cleaned data as CSV string (capped at 5 MB to stay within JSONB)
    buf = io.StringIO()
    df.to_csv(buf, index=False)
    csv_str = buf.getvalue()
    if len(csv_str) > 5 * 1024 * 1024:
        csv_str = csv_str[:5 * 1024 * 1024]
    payload = {
        "tool": "cleaner",
        "name": name,
        "program_id": program_id,
        "data": {
            "csv_content": csv_str,
            "row_count": len(df),
            "col_count": len(df.columns),
            "columns": list(df.columns),
            "filename": DATA.get("filename") or name,
        },
    }
    try:
        resp = _requests.post(f"{fg_base_url}/api/v1/tool-projects/", json=payload,
                              headers={"Authorization": f"Bearer {token}"}, timeout=30)
        return jsonify(resp.json()), resp.status_code
    except Exception as e:
        return jsonify(error=str(e)), 502


@app.route("/api/fg/user-projects/save", methods=["POST"])
def proxy_save_fg_project():
    body = request.json or {}
    fg_base_url = body.get("fg_base_url", "").rstrip("/")
    token = body.get("token", "")
    if not fg_base_url or not token:
        return jsonify(error="fg_base_url and token required"), 400
    payload = {
        "tool": "cleaner",
        "name": body.get("name", "Untitled"),
        "program_id": body.get("program_id"),
        "data": body.get("data", {}),
    }
    try:
        resp = _requests.post(f"{fg_base_url}/api/v1/tool-projects/", json=payload,
                              headers={"Authorization": f"Bearer {token}"}, timeout=30)
        return jsonify(resp.json()), resp.status_code
    except Exception as e:
        return jsonify(error=str(e)), 502


@app.route("/api/load-from-fg", methods=["POST"])
def load_from_fg():
    """Fetch program submissions from FieldGovern; supports optional questionnaire_id filter."""
    body = request.json or {}
    fg_base_url = body.get("fg_base_url", "").rstrip("/")
    program_id = body.get("program_id", "")
    token = body.get("token", "")
    questionnaire_id = body.get("questionnaire_id", "")
    if not fg_base_url or not program_id or not token:
        return jsonify(error="fg_base_url, program_id, and token are required"), 400

    url = f"{fg_base_url}/api/v1/fg/programs/{program_id}/export.xlsx"
    if questionnaire_id:
        url += f"?questionnaire_id={questionnaire_id}"
    try:
        resp = _requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=120)
        if resp.status_code != 200:
            return jsonify(error=f"FieldGovern returned {resp.status_code}"), 502
    except _requests.RequestException as e:
        return jsonify(error=f"Could not reach FieldGovern: {e}"), 502

    filename = f"fg_program_{program_id}.xlsx"
    copy_path = COPIES_DIR / f"COPY_{filename}"
    copy_path.write_bytes(resp.content)

    try:
        xls = pd.ExcelFile(copy_path, engine="openpyxl")
        df = pd.read_excel(copy_path, sheet_name=xls.sheet_names[0], engine="openpyxl",
                           dtype=str, keep_default_na=False, na_values=[""])
        DATA["sheet_name"] = xls.sheet_names[0]
    except Exception as e:
        copy_path.unlink(missing_ok=True)
        return jsonify(error=f"Parse error: {e}"), 400

    df.columns = df.columns.astype(str)
    DATA["df"] = df
    DATA["original_df"] = df.copy()
    DATA["filename"] = filename
    DATA["copy_path"] = str(copy_path)
    DATA["history"] = []
    DATA["column_types"] = {}
    DATA["active_filters"] = {}
    _save_state()
    mark_state_dirty()

    return jsonify(ok=True, filename=filename, rows=len(df), cols=len(df.columns),
                   columns=df.columns.tolist())


# ── Main ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    _load_state()  # Try to restore session on startup
    PORT = int(os.environ.get("PORT", 5050))
    print(f"\n  Data Cleaner running at http://localhost:{PORT}\n")
    app.run(debug=False, use_reloader=False, host="0.0.0.0", port=PORT)
